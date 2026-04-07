"""
PL (Profit & Loss) Generator for FL Cosmetics.
Generates P&L statement from categorized bank transactions + sales report.

PL methodology (from "Гайд P&L_с правками.docx"):
- Выручка = ОП из формы 9 (sales report)
- НДС = Выручка / 1.16
- ОП без НДС = Выручка - НДС
- Себестоимость = из Sales Report (Selfcost USD × курс)
- Валовая прибыль = ОП без НДС - Себестоимость (~70%)
- Расходы склада: аренда=72,000 MXN, расходные=36 MXN×заказы, ФОТ из выписки
- Коммерческие: транспортные=10.4% от ОП без НДС, эквайринг=2.9%
- Комиссии лидерам: ОС + КвБ из выписки
"""

import json
import os
import sys
from collections import defaultdict
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from config import OUTPUT_DIR, SALES_REPORT_FILE

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
except ImportError:
    print("ERROR: openpyxl not installed.")
    sys.exit(1)


# PL Constants (from methodology guide)
VAT_RATE = 0.16  # НДС 16% Mexico
RENT_MONTHLY = 72000  # Аренда склада MXN/month
PACKAGING_PER_ORDER = 36  # Упаковка MXN/заказ
DELIVERY_PCT = 0.104  # Доставка 10.4% от ОП без НДС
ACQUIRING_PCT = 0.029  # Эквайринг 2.9%
WRITEOFF_PCT = 0.01  # Списание 1% от ОП без НДС


def parse_sales_report(filepath=None):
    """Parse sales report for revenue and selfcost data.

    Returns dict with:
    - total_revenue_mxn: total sales in MXN
    - total_selfcost_usd: total selfcost in USD
    - order_count: number of unique orders
    - by_channel: {channel: amount} from TOTAL sheet
    - monthly_data: {month: {revenue, selfcost, orders}} if available
    """
    filepath = filepath or SALES_REPORT_FILE

    if not os.path.exists(filepath):
        print(f"WARNING: Sales report not found: {filepath}")
        return None

    wb = openpyxl.load_workbook(filepath, data_only=True)
    result = {
        "total_revenue_mxn": 0,
        "total_selfcost_usd": 0,
        "order_count": 0,
        "by_channel": {},
    }

    # TOTAL sheet — channel breakdown
    if "TOTAL" in wb.sheetnames:
        ws = wb["TOTAL"]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            if row and row[0]:
                name = str(row[0]).strip()
                val = row[1] if len(row) > 1 else None
                if val and isinstance(val, (int, float)):
                    result["by_channel"][name] = float(val)

    # Main data sheet — individual order lines
    # Columns: c1=Artículo, c3=Nombre, c4=Nº de Orden, c12=Cantidad,
    #          c15=Total(MXN), c16=Selfcost(USD)
    sheet_name = "Sales report" if "Sales report" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    orders = set()
    total_price = 0
    total_selfcost = 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row is None or len(row) < 16:
            continue

        order_id = row[3] if len(row) > 3 else None  # c4 = Nº de Orden
        price_mxn = row[14] if len(row) > 14 else None  # c15 = Total MXN
        selfcost_usd = row[15] if len(row) > 15 else None  # c16 = Selfcost USD

        if order_id:
            orders.add(str(order_id))

        if isinstance(price_mxn, (int, float)):
            total_price += float(price_mxn)
        if isinstance(selfcost_usd, (int, float)):
            total_selfcost += float(selfcost_usd)

    result["total_revenue_mxn"] = total_price
    result["total_selfcost_usd"] = total_selfcost
    result["order_count"] = len(orders)

    wb.close()
    return result


def generate_pl_data(categorized_transactions, sales_data=None, usd_mxn_rate=None):
    """Generate P&L data from categorized transactions and sales report.

    Args:
        categorized_transactions: list of categorized transaction dicts
        sales_data: output of parse_sales_report()
        usd_mxn_rate: USD/MXN exchange rate for selfcost conversion

    Returns dict of PL line items.
    """
    from categorizer import normalize_category

    # Aggregate bank data by category
    by_cat = defaultdict(lambda: {"cargo": 0.0, "abono": 0.0, "count": 0})
    monthly_abono = defaultdict(float)

    for txn in categorized_transactions:
        cat = txn.get("predicted_category", "UNKNOWN")
        by_cat[cat]["cargo"] += txn.get("cargo", 0) or 0
        by_cat[cat]["abono"] += txn.get("abono", 0) or 0
        by_cat[cat]["count"] += 1
        monthly_abono[txn["month"]] += txn.get("abono", 0) or 0

    num_months = len(monthly_abono)

    # Revenue from bank statement (abono marked as ОП)
    revenue_bank = by_cat["ОП"]["abono"]

    # Revenue from sales report (if available)
    revenue_sales = sales_data["total_revenue_mxn"] if sales_data else revenue_bank

    # Use bank statement revenue as primary (it's actual cash received)
    revenue = revenue_bank

    # PL calculations per methodology
    vat = revenue / (1 + VAT_RATE) * VAT_RATE  # НДС
    revenue_ex_vat = revenue - vat  # ОП без НДС

    # Себестоимость
    # usd_mxn_rate can be a dict {month: rate} or a single float
    if isinstance(usd_mxn_rate, dict):
        # Use average of all available rates
        avg_rate = sum(usd_mxn_rate.values()) / len(usd_mxn_rate) if usd_mxn_rate else 20.5
    else:
        avg_rate = usd_mxn_rate or 20.5

    if sales_data and sales_data["total_selfcost_usd"] > 0:
        selfcost = sales_data["total_selfcost_usd"] * avg_rate
    else:
        # Approximate: ~30% of revenue ex VAT (typical for Faberlic)
        selfcost = revenue_ex_vat * 0.30

    gross_profit = revenue_ex_vat - selfcost
    gross_margin = gross_profit / revenue_ex_vat if revenue_ex_vat > 0 else 0

    # Order count
    order_count = sales_data["order_count"] if sales_data else 0

    # Расходы склада
    rent = RENT_MONTHLY * num_months
    packaging = PACKAGING_PER_ORDER * order_count if order_count > 0 else by_cat.get("Упаковка", {}).get("cargo", 0)
    warehouse_salary = by_cat["ЗП склад"]["cargo"]
    warehouse_utilities = by_cat["Коммунальные платежи"]["cargo"]
    warehouse_maintenance = by_cat["Хозяйственные расходы"]["cargo"]
    warehouse_materials = by_cat["Расходные материалы"]["cargo"]
    warehouse_internet = by_cat["Связь"]["cargo"]
    total_warehouse = rent + packaging + warehouse_salary + warehouse_utilities + warehouse_maintenance + warehouse_materials + warehouse_internet

    # Коммерческие расходы
    transport = by_cat["Транспортные в регион"]["cargo"]
    # transport_formula = revenue_ex_vat * DELIVERY_PCT  # formula-based
    acquiring = revenue_ex_vat * ACQUIRING_PCT
    agent_commission = 0  # агентское вознаграждение
    total_commercial_variable = transport + acquiring + agent_commission

    # ОС лидерам + КвБ
    leader_os = by_cat["ОС лидерам"]["cargo"]
    leader_kv = 0  # КвБ — separate if available
    total_leader = leader_os + leader_kv

    # Постоянные расходы
    office_salary = by_cat["ЗП офис"]["cargo"]
    stimulation = by_cat["Стимулирование продаж"]["cargo"]
    accounting = by_cat["Бухгалтерские услуги"]["cargo"]
    bank_services = by_cat["Услуги банка"]["cargo"]
    other_expenses = by_cat["Прочие расходы"]["cargo"]
    catalog = by_cat["Реклама каталог"]["cargo"]
    writeoffs = revenue_ex_vat * WRITEOFF_PCT
    customs = by_cat["Услуги по таможенному оформлению"]["cargo"]

    total_fixed = office_salary + stimulation + accounting + bank_services + other_expenses + catalog + writeoffs + customs

    # Налоги
    tax_fot = by_cat["Налог на сотрудников"]["cargo"]
    tax_leaders = by_cat["Налог на выплаты лидерам"]["cargo"]
    tax_vat = by_cat["НДС"]["cargo"]
    total_tax = tax_fot + tax_leaders + tax_vat

    # Итого операционные расходы
    total_opex = total_warehouse + total_commercial_variable + total_leader + total_fixed + total_tax

    # Операционная прибыль
    operating_profit = gross_profit - total_opex

    # Build result
    pl = {
        "period": f"{num_months} months",
        "revenue": revenue,
        "vat": vat,
        "revenue_ex_vat": revenue_ex_vat,
        "selfcost": selfcost,
        "gross_profit": gross_profit,
        "gross_margin": gross_margin,

        # Расходы склада
        "warehouse": {
            "rent": rent,
            "packaging": packaging,
            "salary": warehouse_salary,
            "utilities": warehouse_utilities,
            "maintenance": warehouse_maintenance,
            "materials": warehouse_materials,
            "internet": warehouse_internet,
            "total": total_warehouse,
        },

        # Коммерческие переменные
        "commercial": {
            "transport": transport,
            "acquiring": acquiring,
            "agent": agent_commission,
            "total": total_commercial_variable,
        },

        # Лидеры
        "leaders": {
            "os": leader_os,
            "kv": leader_kv,
            "total": total_leader,
        },

        # Постоянные
        "fixed": {
            "office_salary": office_salary,
            "stimulation": stimulation,
            "accounting": accounting,
            "bank_services": bank_services,
            "other": other_expenses,
            "catalog": catalog,
            "writeoffs": writeoffs,
            "customs": customs,
            "total": total_fixed,
        },

        # Налоги
        "taxes": {
            "fot": tax_fot,
            "leaders": tax_leaders,
            "vat": tax_vat,
            "total": total_tax,
        },

        "total_opex": total_opex,
        "operating_profit": operating_profit,
        "operating_margin": operating_profit / revenue_ex_vat if revenue_ex_vat > 0 else 0,
    }

    return pl


def write_pl_excel(pl_data, output_path=None):
    """Write P&L to Excel file."""
    output_path = output_path or os.path.join(OUTPUT_DIR, "PL_факт.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PL факт"

    # Styles
    header_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=11)
    subsection_font = Font(bold=True, size=10)
    number_format = '#,##0.00'
    pct_format = '0.0%'
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    section_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    rev = pl_data["revenue_ex_vat"]

    def pct(val):
        return val / rev if rev > 0 else 0

    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 12

    # Header
    ws.cell(1, 1, "FL COSMETICS MEXICO — P&L (факт)").font = header_font
    ws.cell(2, 1, f"Период: {pl_data['period']}")

    # Column headers
    for col, header in [(1, "Статья"), (2, "Факт, MXN"), (3, "% от ОП б/НДС")]:
        cell = ws.cell(4, col, header)
        cell.fill = header_fill
        cell.font = header_font_white
        cell.border = thin_border

    def add_row(r, label, value, is_pct=True, level=0, is_section=False, is_total=False):
        indent = "  " * level
        c1 = ws.cell(r, 1, f"{indent}{label}")
        c2 = ws.cell(r, 2, value)
        c2.number_format = number_format

        if is_pct and rev > 0:
            c3 = ws.cell(r, 3, value / rev)
            c3.number_format = pct_format

        c1.border = thin_border
        c2.border = thin_border
        ws.cell(r, 3).border = thin_border

        if is_section:
            c1.font = section_font
            c1.fill = section_fill
        if is_total:
            c1.font = subsection_font
            c1.fill = green_fill
            c2.font = subsection_font

    r = 5
    add_row(r, "Объем продаж (с НДС)", pl_data["revenue"], False, 0, True); r += 1
    add_row(r, "НДС", pl_data["vat"], False, 1); r += 1
    add_row(r, "ОП без НДС", pl_data["revenue_ex_vat"], True, 0, True); r += 1
    add_row(r, "Себестоимость", pl_data["selfcost"], True, 0); r += 1
    add_row(r, "Валовая прибыль", pl_data["gross_profit"], True, 0, False, True); r += 1

    r += 1
    add_row(r, "ОПЕРАЦИОННЫЕ РАСХОДЫ", 0, False, 0, True); r += 1

    # Переменные расходы
    add_row(r, "Расходы склада", 0, False, 0, True); r += 1
    w = pl_data["warehouse"]
    add_row(r, "аренда склада", w["rent"], True, 1); r += 1
    add_row(r, "расходные материалы (упаковка)", w["packaging"], True, 1); r += 1
    add_row(r, "зарплата и ФОТ склад", w["salary"], True, 1); r += 1
    add_row(r, "коммунальные расходы", w["utilities"], True, 1); r += 1
    add_row(r, "хозяйственные расходы", w["maintenance"], True, 1); r += 1
    add_row(r, "расходные материалы", w["materials"], True, 1); r += 1
    add_row(r, "интернет и связь", w["internet"], True, 1); r += 1
    add_row(r, "Итого расходы склада", w["total"], True, 0, False, True); r += 1

    r += 1
    add_row(r, "Коммерческие расходы", 0, False, 0, True); r += 1
    c = pl_data["commercial"]
    add_row(r, "транспортные (в регионы)", c["transport"], True, 1); r += 1
    add_row(r, "услуги за терминальную оплату (эквайринг 2.9%)", c["acquiring"], True, 1); r += 1
    add_row(r, "Итого коммерческие", c["total"], True, 0, False, True); r += 1

    r += 1
    add_row(r, "Комиссии лидерам", 0, False, 0, True); r += 1
    l = pl_data["leaders"]
    add_row(r, "объемная скидка (ОС)", l["os"], True, 1); r += 1
    add_row(r, "квалификационный бонус", l["kv"], True, 1); r += 1
    add_row(r, "Итого комиссии лидерам", l["total"], True, 0, False, True); r += 1

    r += 1
    add_row(r, "Постоянные расходы", 0, False, 0, True); r += 1
    f = pl_data["fixed"]
    add_row(r, "заработная плата офис", f["office_salary"], True, 1); r += 1
    add_row(r, "стимулирование продаж", f["stimulation"], True, 1); r += 1
    add_row(r, "аудит, бух.учет", f["accounting"], True, 1); r += 1
    add_row(r, "услуги банка (РКО)", f["bank_services"], True, 1); r += 1
    add_row(r, "прочие расходы", f["other"], True, 1); r += 1
    add_row(r, "реклама каталог", f["catalog"], True, 1); r += 1
    add_row(r, "списание продукции (1%)", f["writeoffs"], True, 1); r += 1
    add_row(r, "таможенное оформление", f["customs"], True, 1); r += 1
    add_row(r, "Итого постоянные расходы", f["total"], True, 0, False, True); r += 1

    r += 1
    add_row(r, "Налоги", 0, False, 0, True); r += 1
    t = pl_data["taxes"]
    add_row(r, "налоги ФОТ", t["fot"], True, 1); r += 1
    add_row(r, "налоги за выплаты лидерам", t["leaders"], True, 1); r += 1
    add_row(r, "НДС уплаченный", t["vat"], True, 1); r += 1
    add_row(r, "Итого налоги", t["total"], True, 0, False, True); r += 1

    r += 1
    add_row(r, "ИТОГО ОПЕРАЦИОННЫЕ РАСХОДЫ", pl_data["total_opex"], True, 0, True); r += 1
    r += 1
    cell = ws.cell(r, 1, "ОПЕРАЦИОННАЯ ПРИБЫЛЬ")
    cell.font = Font(bold=True, size=12)
    cell.fill = green_fill
    cell.border = thin_border
    c2 = ws.cell(r, 2, pl_data["operating_profit"])
    c2.number_format = number_format
    c2.font = Font(bold=True, size=12)
    c2.border = thin_border
    c3 = ws.cell(r, 3, pl_data["operating_margin"])
    c3.number_format = pct_format
    c3.font = Font(bold=True)
    c3.border = thin_border

    wb.save(output_path)
    print(f"PL saved to {output_path}")
    return output_path


def print_pl_summary(pl):
    """Print P&L summary to console."""
    rev = pl["revenue_ex_vat"]

    def pct(val):
        return f"{val/rev*100:.1f}%" if rev > 0 else "0%"

    print(f"\n{'='*60}")
    print(f"FL COSMETICS MEXICO — P&L SUMMARY")
    print(f"Period: {pl['period']}")
    print(f"{'='*60}")
    print(f"  Объем продаж (с НДС)         {pl['revenue']:>15,.0f}")
    print(f"  НДС                          {pl['vat']:>15,.0f}")
    print(f"  ОП без НДС                   {pl['revenue_ex_vat']:>15,.0f}  100.0%")
    print(f"  Себестоимость                 {pl['selfcost']:>15,.0f}  {pct(pl['selfcost'])}")
    print(f"  ВАЛОВАЯ ПРИБЫЛЬ              {pl['gross_profit']:>15,.0f}  {pct(pl['gross_profit'])}")
    print(f"")
    print(f"  Расходы склада               {pl['warehouse']['total']:>15,.0f}  {pct(pl['warehouse']['total'])}")
    print(f"    аренда                      {pl['warehouse']['rent']:>15,.0f}")
    print(f"    упаковка                    {pl['warehouse']['packaging']:>15,.0f}")
    print(f"    ЗП склад                    {pl['warehouse']['salary']:>15,.0f}")
    print(f"    коммунальные                {pl['warehouse']['utilities']:>15,.0f}")
    print(f"  Коммерческие расходы          {pl['commercial']['total']:>15,.0f}  {pct(pl['commercial']['total'])}")
    print(f"    транспортные                {pl['commercial']['transport']:>15,.0f}")
    print(f"    эквайринг (2.9%)            {pl['commercial']['acquiring']:>15,.0f}")
    print(f"  Комиссии лидерам              {pl['leaders']['total']:>15,.0f}  {pct(pl['leaders']['total'])}")
    print(f"  Постоянные расходы            {pl['fixed']['total']:>15,.0f}  {pct(pl['fixed']['total'])}")
    print(f"    ЗП офис                     {pl['fixed']['office_salary']:>15,.0f}")
    print(f"    стимулирование              {pl['fixed']['stimulation']:>15,.0f}")
    print(f"    бух.услуги                  {pl['fixed']['accounting']:>15,.0f}")
    print(f"    реклама каталог             {pl['fixed']['catalog']:>15,.0f}")
    print(f"    таможня                     {pl['fixed']['customs']:>15,.0f}")
    print(f"  Налоги                        {pl['taxes']['total']:>15,.0f}  {pct(pl['taxes']['total'])}")
    print(f"")
    print(f"  ИТОГО OPEX                   {pl['total_opex']:>15,.0f}  {pct(pl['total_opex'])}")
    print(f"  ОПЕРАЦИОННАЯ ПРИБЫЛЬ         {pl['operating_profit']:>15,.0f}  {pct(pl['operating_profit'])}")
    print(f"{'='*60}")


if __name__ == "__main__":
    from parser import parse_bank_statement
    from categorizer import (
        build_dictionary_from_bank, build_dictionary_from_william,
        merge_dictionaries, categorize_all
    )
    from config import PAYMENTS_FILE

    print("=" * 60)
    print("FL Cosmetics — P&L Generator")
    print("=" * 60)

    # 1. Parse & categorize
    print("\n1. Parsing and categorizing transactions...")
    txns = parse_bank_statement()
    bank_dict = build_dictionary_from_bank(txns)
    william_dict = build_dictionary_from_william(PAYMENTS_FILE)
    merged = merge_dictionaries(bank_dict, william_dict)
    categorized = categorize_all(txns, merged)

    # 2. Parse sales report
    print("\n2. Parsing sales report...")
    sales = parse_sales_report()
    if sales:
        print(f"   Revenue: {sales['total_revenue_mxn']:,.0f} MXN")
        print(f"   Selfcost: {sales['total_selfcost_usd']:,.0f} USD")
        print(f"   Orders: {sales['order_count']}")
        print(f"   Channels: {sales['by_channel']}")

    # 3. Generate PL
    print("\n3. Generating P&L...")
    pl = generate_pl_data(categorized, sales)

    # 4. Print
    print_pl_summary(pl)

    # 5. Write Excel
    print("\n4. Writing Excel file...")
    path = write_pl_excel(pl)
    print(f"\nDone! File: {path}")
