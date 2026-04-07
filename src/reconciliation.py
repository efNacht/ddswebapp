"""
Provider Reconciliation for FL Cosmetics.
Matches payment provider reports (Cards/Bank/Cash/MercadoPago) against
Santander bank statement to find discrepancies.

Provider data comes from Sales Report (DOC-20260204-WA0002).
Bank data comes from categorized bank transactions (ОП = incoming payments).
"""

import os
import sys
from collections import defaultdict
from datetime import datetime, timedelta

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from config import SALES_REPORT_FILE, OUTPUT_DIR

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
except ImportError:
    print("ERROR: openpyxl not installed.")
    sys.exit(1)


def parse_provider_data(filepath=None):
    """Parse provider payment details from sales report.

    Returns dict with keys: cards, bank, cash, mercado.
    Each is a list of dicts with: id, date, amount, description, status.
    Also returns totals dict.
    """
    filepath = filepath or SALES_REPORT_FILE

    if not os.path.exists(filepath):
        print(f"ERROR: Sales report not found: {filepath}")
        return None, None

    wb = openpyxl.load_workbook(filepath, data_only=True)
    providers = {}

    # TOTAL sheet — summary
    totals = {}
    if "TOTAL" in wb.sheetnames:
        ws = wb["TOTAL"]
        for r in range(1, ws.max_row + 1):
            name = ws.cell(r, 1).value
            val = ws.cell(r, 2).value
            if name and val and isinstance(val, (int, float)):
                totals[str(name).strip().upper()] = float(val)

    # CARDS — OpenPay card payments
    if "CARDS" in wb.sheetnames:
        ws = wb["CARDS"]
        cards = []
        for r in range(2, ws.max_row + 1):
            txn_id = ws.cell(r, 1).value  # transaction_id
            desc = ws.cell(r, 3).value  # description (order number)
            date = ws.cell(r, 4).value  # creation_date
            status = ws.cell(r, 6).value  # status
            amount = ws.cell(r, 18).value  # amount (column 18)

            if not amount or not isinstance(amount, (int, float)):
                continue

            cards.append({
                "id": str(txn_id or ""),
                "date": date if isinstance(date, datetime) else None,
                "amount": float(amount),
                "description": str(desc or ""),
                "status": str(status or ""),
                "provider": "CARDS",
            })
        providers["cards"] = cards

    # BANK — bank transfer payments via OpenPay
    if "BANK" in wb.sheetnames:
        ws = wb["BANK"]
        bank = []
        for r in range(2, ws.max_row + 1):
            txn_id = ws.cell(r, 2).value  # transaction_id
            date = ws.cell(r, 4).value  # creation_date
            amount = ws.cell(r, 6).value  # amount
            status = ws.cell(r, 8).value  # status
            desc = ws.cell(r, 9).value  # description

            if not amount or not isinstance(amount, (int, float)):
                continue

            bank.append({
                "id": str(txn_id or ""),
                "date": date if isinstance(date, datetime) else None,
                "amount": float(amount),
                "description": str(desc or ""),
                "status": str(status or ""),
                "provider": "BANK",
            })
        providers["bank"] = bank

    # CASH — store payments (Farmacias, Walmart)
    if "CASH" in wb.sheetnames:
        ws = wb["CASH"]
        cash = []
        for r in range(2, ws.max_row + 1):
            txn_id = ws.cell(r, 1).value
            desc = ws.cell(r, 3).value
            date = ws.cell(r, 4).value
            status = ws.cell(r, 6).value
            store = ws.cell(r, 8).value  # store_name

            # CASH: amount is in column 16
            amount = ws.cell(r, 16).value
            if not isinstance(amount, (int, float)):
                amount = None

            if amount is None:
                continue

            cash.append({
                "id": str(txn_id or ""),
                "date": date if isinstance(date, datetime) else None,
                "amount": amount,
                "description": f"{desc or ''} @ {store or ''}",
                "status": str(status or ""),
                "provider": "CASH",
                "store": str(store or ""),
            })
        providers["cash"] = cash

    wb.close()
    return providers, totals


def get_bank_income_by_type(categorized_transactions):
    """Extract income transactions from bank statement grouped by likely provider.

    Identifies provider type from concept/description:
    - Cards (OpenPay): concept contains "PAGO XYS" or specific patterns
    - Bank transfers: concept contains SPEI references for ОП
    - Cash: concept mentions "deposito efectivo" or store patterns
    - MercadoPago: concept contains "MERCADO*PAGO" or "MP"

    Returns dict with: cards_total, bank_total, cash_total, mercado_total, other_total.
    """
    income = defaultdict(float)
    income_txns = defaultdict(list)

    for txn in categorized_transactions:
        if txn.get("predicted_category") != "ОП":
            continue

        abono = txn.get("abono", 0) or 0
        if abono <= 0:
            continue

        desc = (txn.get("description", "") or "").upper()
        concept = (txn.get("concept", "") or "").upper()

        # Classify by provider
        if "MERCADO" in concept or "MERCADO" in desc:
            provider = "MERCADO"
        elif "PAGO XYS" in concept or "OPEN XYS" in concept:
            provider = "OPENPAY"  # OpenPay settlements (Cards+Bank+Cash combined)
        elif "DEPOSITO EFECTIVO" in desc or "DEPOSITO" in desc:
            provider = "CASH"
        else:
            provider = "OTHER"

        income[provider] += abono
        income_txns[provider].append(txn)

    return dict(income), dict(income_txns)


def reconcile(providers, totals, categorized_transactions):
    """Reconcile provider totals against bank statement.

    Key insight: OpenPay sends consolidated SPEI settlements combining
    Cards + Bank + Cash into single transfers ("PAGO XYS..." pattern).
    We can't split them on the bank side, so we compare:
    - OpenPay combined (Cards+Bank+Cash) vs bank OPENPAY total
    - MercadoPago vs bank MERCADO total

    Note: provider report covers ~1 period, bank covers 14 months.
    """
    bank_income, bank_txns = get_bank_income_by_type(categorized_transactions)

    total_bank_op = sum(bank_income.values())

    # Provider totals from sales report
    provider_totals = {
        "CARDS": totals.get("CARDS", 0),
        "BANK": totals.get("BANK", 0),
        "CASH": totals.get("CASH", 0),
        "MERCADO": totals.get("MERCADO", 0),
    }
    provider_grand = sum(provider_totals.values())

    # OpenPay combined = Cards + Bank + Cash (they send consolidated SPEI)
    openpay_provider = provider_totals["CARDS"] + provider_totals["BANK"] + provider_totals["CASH"]

    # Provider detail sums
    detail_sums = {}
    for channel, txns in providers.items():
        detail_sums[channel.upper()] = sum(t["amount"] for t in txns)

    # Bank statement income by detected provider
    bank_by_provider = {
        "OPENPAY": bank_income.get("OPENPAY", 0),
        "MERCADO": bank_income.get("MERCADO", 0),
        "CASH": bank_income.get("CASH", 0),
        "OTHER": bank_income.get("OTHER", 0),
    }

    # Transaction counts
    bank_counts = {k: len(v) for k, v in bank_txns.items()}

    # Monthly breakdown for bank income
    monthly_bank = defaultdict(lambda: defaultdict(float))
    for provider_key, txn_list in bank_txns.items():
        for txn in txn_list:
            month = txn.get("month", "Unknown")
            monthly_bank[month][provider_key] += txn.get("abono", 0) or 0

    report = {
        "provider_totals": provider_totals,
        "provider_detail_sums": detail_sums,
        "openpay_provider_combined": openpay_provider,
        "bank_income_total": total_bank_op,
        "bank_by_provider": bank_by_provider,
        "bank_counts": bank_counts,
        "provider_grand_total": provider_grand,
        "monthly_bank": dict(monthly_bank),
        "checks": [],
    }

    # Check 1: Provider internal consistency (TOTAL sheet vs detail rows)
    for channel in ["CARDS", "BANK", "CASH"]:
        total_val = provider_totals.get(channel, 0)
        detail_val = detail_sums.get(channel, 0)
        diff = detail_val - total_val
        report["checks"].append({
            "check": f"{channel}: TOTAL vs detail rows",
            "expected": total_val,
            "actual": detail_val,
            "diff": diff,
            "diff_pct": (diff / total_val * 100) if total_val > 0 else 0,
            "status": "OK" if abs(diff) < 1 else "MISMATCH",
        })

    # Check 2: OpenPay combined (provider) vs OPENPAY (bank) — period mismatch expected
    report["checks"].append({
        "check": "OpenPay combined (provider ~1 period) vs bank OPENPAY (14 months)",
        "expected": openpay_provider,
        "actual": bank_by_provider["OPENPAY"],
        "diff": bank_by_provider["OPENPAY"] - openpay_provider,
        "diff_pct": ((bank_by_provider["OPENPAY"] - openpay_provider) / openpay_provider * 100) if openpay_provider > 0 else 0,
        "status": "PERIOD_MISMATCH",
    })

    # Check 3: MercadoPago (provider ~1 period) vs bank MERCADO (14 months)
    report["checks"].append({
        "check": "MercadoPago (provider ~1 period) vs bank MERCADO (14 months)",
        "expected": provider_totals["MERCADO"],
        "actual": bank_by_provider["MERCADO"],
        "diff": bank_by_provider["MERCADO"] - provider_totals["MERCADO"],
        "diff_pct": ((bank_by_provider["MERCADO"] - provider_totals["MERCADO"]) / provider_totals["MERCADO"] * 100) if provider_totals["MERCADO"] > 0 else 0,
        "status": "PERIOD_MISMATCH",
    })

    return report


def write_reconciliation_excel(report, output_path=None):
    """Write reconciliation report to Excel."""
    output_path = output_path or os.path.join(OUTPUT_DIR, "Сверка_провайдеров.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Сверка"

    # Styles
    header_font = Font(bold=True, size=12)
    section_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    red_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    num_fmt = '#,##0.00'
    pct_fmt = '0.0%'

    ws.column_dimensions['A'].width = 55
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14

    # Title
    ws.cell(1, 1, "FL COSMETICS — Сверка провайдеров").font = header_font

    # --- Section 1: Provider report breakdown ---
    r = 3
    ws.cell(r, 1, "1. Отчёт провайдеров (Sales Report, ~1 период)").font = section_font
    r += 1
    for col, hdr in enumerate(["Канал", "TOTAL (MXN)", "Детализация (MXN)", "Разница", "Статус"], 1):
        cell = ws.cell(r, col, hdr)
        cell.fill = header_fill
        cell.font = header_font_white
        cell.border = thin_border
    r += 1
    for ch in ["CARDS", "BANK", "CASH"]:
        ws.cell(r, 1, ch).border = thin_border
        total_val = report["provider_totals"].get(ch, 0)
        detail_val = report["provider_detail_sums"].get(ch, 0)
        diff = detail_val - total_val
        ws.cell(r, 2, total_val).number_format = num_fmt
        ws.cell(r, 2).border = thin_border
        ws.cell(r, 3, detail_val).number_format = num_fmt
        ws.cell(r, 3).border = thin_border
        ws.cell(r, 4, diff).number_format = num_fmt
        ws.cell(r, 4).border = thin_border
        status = "OK" if abs(diff) < 1 else "MISMATCH"
        ws.cell(r, 5, status).border = thin_border
        if status == "OK":
            ws.cell(r, 5).fill = green_fill
        else:
            ws.cell(r, 5).fill = red_fill
        r += 1

    # OpenPay combined
    ws.cell(r, 1, "OpenPay итого (Cards+Bank+Cash)").font = section_font
    ws.cell(r, 1).border = thin_border
    ws.cell(r, 2, report["openpay_provider_combined"]).number_format = num_fmt
    ws.cell(r, 2).font = section_font
    ws.cell(r, 2).border = thin_border
    r += 1
    ws.cell(r, 1, "MercadoPago").border = thin_border
    ws.cell(r, 2, report["provider_totals"]["MERCADO"]).number_format = num_fmt
    ws.cell(r, 2).border = thin_border
    r += 1
    ws.cell(r, 1, "ИТОГО провайдеры").font = section_font
    ws.cell(r, 1).border = thin_border
    ws.cell(r, 2, report["provider_grand_total"]).number_format = num_fmt
    ws.cell(r, 2).font = section_font
    ws.cell(r, 2).border = thin_border

    # --- Section 2: Bank income breakdown ---
    r += 2
    ws.cell(r, 1, "2. Выписка банка — доход (ОП) за 14 месяцев").font = section_font
    r += 1
    for col, hdr in enumerate(["Источник", "Сумма (MXN)", "Кол-во транзакций"], 1):
        cell = ws.cell(r, col, hdr)
        cell.fill = header_fill
        cell.font = header_font_white
        cell.border = thin_border
    r += 1
    for src in ["OPENPAY", "MERCADO", "CASH", "OTHER"]:
        ws.cell(r, 1, src).border = thin_border
        ws.cell(r, 2, report["bank_by_provider"].get(src, 0)).number_format = num_fmt
        ws.cell(r, 2).border = thin_border
        ws.cell(r, 3, report["bank_counts"].get(src, 0)).border = thin_border
        r += 1
    ws.cell(r, 1, "ИТОГО банк ОП").font = section_font
    ws.cell(r, 1).border = thin_border
    ws.cell(r, 2, report["bank_income_total"]).number_format = num_fmt
    ws.cell(r, 2).font = section_font
    ws.cell(r, 2).border = thin_border

    # --- Section 3: Checks ---
    r += 2
    ws.cell(r, 1, "3. Проверки").font = section_font
    r += 1
    for col, hdr in enumerate(["Проверка", "Ожидаемо", "Факт", "Разница", "%", "Статус"], 1):
        cell = ws.cell(r, col, hdr)
        cell.fill = header_fill
        cell.font = header_font_white
        cell.border = thin_border
    for chk in report["checks"]:
        r += 1
        ws.cell(r, 1, chk["check"]).border = thin_border
        ws.cell(r, 2, chk["expected"]).number_format = num_fmt
        ws.cell(r, 2).border = thin_border
        ws.cell(r, 3, chk["actual"]).number_format = num_fmt
        ws.cell(r, 3).border = thin_border
        ws.cell(r, 4, chk["diff"]).number_format = num_fmt
        ws.cell(r, 4).border = thin_border
        pct_val = chk["diff_pct"] / 100 if chk["diff_pct"] else 0
        ws.cell(r, 5, pct_val).number_format = pct_fmt
        ws.cell(r, 5).border = thin_border
        ws.cell(r, 6, chk["status"]).border = thin_border
        if chk["status"] == "OK":
            ws.cell(r, 6).fill = green_fill
        elif chk["status"] == "PERIOD_MISMATCH":
            ws.cell(r, 6).fill = yellow_fill
        else:
            ws.cell(r, 6).fill = red_fill

    # --- Section 4: Notes ---
    r += 2
    ws.cell(r, 1, "Примечания:").font = section_font
    r += 1
    ws.cell(r, 1, "* OpenPay объединяет Cards+Bank+Cash в одну SPEI-выплату ('PAGO XYS...')")
    r += 1
    ws.cell(r, 1, "* Отчёт провайдеров покрывает ~1 период, банк — 14 месяцев (Фев 2025 – Мар 2026)")
    r += 1
    ws.cell(r, 1, "* Для точной помесячной сверки нужны помесячные выгрузки от OpenPay и MercadoPago")
    r += 1
    ws.cell(r, 1, "* CASH в банке = депозиты наличными (DEPOSITO EFECTIVO), OTHER = прочие ОП переводы")

    # --- Sheet 2: Monthly breakdown ---
    ws2 = wb.create_sheet("По месяцам")
    ws2.column_dimensions['A'].width = 20

    from dds_generator import MONTH_ORDER
    months = sorted(report["monthly_bank"].keys(), key=lambda m: MONTH_ORDER.get(m, 99))

    ws2.cell(1, 1, "Доход (ОП) по месяцам и провайдерам").font = header_font
    r = 3
    headers = ["Месяц", "OPENPAY", "MERCADO", "CASH", "OTHER", "ИТОГО"]
    for col, hdr in enumerate(headers, 1):
        cell = ws2.cell(r, col, hdr)
        cell.fill = header_fill
        cell.font = header_font_white
        cell.border = thin_border
        if col > 1:
            ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 16

    for month in months:
        r += 1
        ws2.cell(r, 1, month).border = thin_border
        month_data = report["monthly_bank"].get(month, {})
        row_total = 0
        for j, src in enumerate(["OPENPAY", "MERCADO", "CASH", "OTHER"], 2):
            val = month_data.get(src, 0)
            ws2.cell(r, j, val).number_format = num_fmt
            ws2.cell(r, j).border = thin_border
            row_total += val
        ws2.cell(r, 6, row_total).number_format = num_fmt
        ws2.cell(r, 6).border = thin_border
        ws2.cell(r, 6).font = Font(bold=True)

    # Totals row
    r += 1
    ws2.cell(r, 1, "ИТОГО").font = section_font
    ws2.cell(r, 1).border = thin_border
    grand = 0
    for j, src in enumerate(["OPENPAY", "MERCADO", "CASH", "OTHER"], 2):
        val = sum(report["monthly_bank"].get(m, {}).get(src, 0) for m in months)
        ws2.cell(r, j, val).number_format = num_fmt
        ws2.cell(r, j).font = section_font
        ws2.cell(r, j).border = thin_border
        grand += val
    ws2.cell(r, 6, grand).number_format = num_fmt
    ws2.cell(r, 6).font = section_font
    ws2.cell(r, 6).border = thin_border

    wb.save(output_path)
    print(f"Reconciliation saved to {output_path}")
    return output_path


def print_reconciliation(report):
    """Print reconciliation report to console."""
    print(f"\n{'='*70}")
    print(f"FL COSMETICS — СВЕРКА ПРОВАЙДЕРОВ")
    print(f"{'='*70}")

    # Provider report
    print(f"\n--- Отчёт провайдеров (Sales Report, ~1 период) ---")
    print(f"  CARDS:      {report['provider_totals']['CARDS']:>12,.0f} MXN")
    print(f"  BANK:       {report['provider_totals']['BANK']:>12,.0f} MXN")
    print(f"  CASH:       {report['provider_totals']['CASH']:>12,.0f} MXN")
    print(f"  ─────────────────────────────")
    print(f"  OpenPay:    {report['openpay_provider_combined']:>12,.0f} MXN (Cards+Bank+Cash)")
    print(f"  MercadoPago:{report['provider_totals']['MERCADO']:>12,.0f} MXN")
    print(f"  ИТОГО:      {report['provider_grand_total']:>12,.0f} MXN")

    # Bank income
    print(f"\n--- Банк: доход (ОП) за 14 месяцев ---")
    print(f"{'Источник':<20s} {'Сумма':>14s} {'Транзакций':>12s}")
    print("-" * 48)
    for src in ["OPENPAY", "MERCADO", "CASH", "OTHER"]:
        amt = report["bank_by_provider"].get(src, 0)
        cnt = report["bank_counts"].get(src, 0)
        print(f"{src:<20s} {amt:>14,.0f} {cnt:>12d}")
    print("-" * 48)
    print(f"{'ИТОГО':<20s} {report['bank_income_total']:>14,.0f}")

    # Checks
    print(f"\n--- Проверки ---")
    for chk in report["checks"]:
        status = chk["status"]
        marker = " [OK]" if status == "OK" else f" [{status}]"
        print(f"  {chk['check']}")
        print(f"    ожидаемо={chk['expected']:>12,.0f}  факт={chk['actual']:>12,.0f}  Δ={chk['diff']:>12,.0f}{marker}")

    print(f"\n{'='*70}")


if __name__ == "__main__":
    from parser import parse_bank_statement
    from categorizer import (
        build_dictionary_from_bank, build_dictionary_from_william,
        merge_dictionaries, categorize_all
    )
    from config import PAYMENTS_FILE

    print("=" * 60)
    print("FL Cosmetics — Provider Reconciliation")
    print("=" * 60)

    # 1. Parse & categorize bank
    print("\n1. Parsing bank statement...")
    txns = parse_bank_statement()
    bank_dict = build_dictionary_from_bank(txns)
    william_dict = build_dictionary_from_william(PAYMENTS_FILE)
    merged = merge_dictionaries(bank_dict, william_dict)
    categorized = categorize_all(txns, merged)

    # 2. Parse provider data
    print("\n2. Parsing provider reports...")
    providers, totals = parse_provider_data()
    if providers:
        for ch, items in providers.items():
            total = sum(t["amount"] for t in items)
            print(f"   {ch.upper()}: {len(items)} transactions, {total:,.0f} MXN")
    if totals:
        print(f"   TOTAL report: {totals}")

    # 3. Reconcile
    print("\n3. Running reconciliation...")
    report = reconcile(providers, totals, categorized)
    print_reconciliation(report)

    # 4. Write Excel
    print("\n4. Writing Excel...")
    path = write_reconciliation_excel(report)
    print(f"\nDone! File: {path}")
