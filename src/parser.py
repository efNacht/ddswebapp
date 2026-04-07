"""
Bank Statement Parser for FL Cosmetics (Santander Mexico)
Parses "Выписка Сантадер.xlsx" — monthly sheets with transaction data.
"""

import re
import sys
import os
from datetime import datetime, time as dt_time

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from config import BANK_STATEMENT_FILE, BANK_SKIP_SHEETS, BANK_COLUMNS

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip3 install openpyxl")
    sys.exit(1)


def parse_amount(val):
    """Parse monetary amount from various formats.

    Handles:
    - None, 0, 0.0 → 0.0
    - Numeric floats → float
    - Strings like "10 000,00" → 10000.00
    - Strings like "1,466,129.58" → 1466129.58
    - Strings like "  1,922,110.97 " → 1922110.97
    """
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)

    s = str(val).strip()
    if not s or s == '0':
        return 0.0

    # Remove currency symbols and whitespace
    s = s.replace('$', '').replace('MXN', '').strip()

    # Detect format: "10 000,00" (space as thousands, comma as decimal)
    if re.match(r'^[\d\s]+,\d{2}$', s.replace(' ', '')):
        # Remove spaces, replace comma with dot
        s = s.replace(' ', '').replace(',', '.')
        return float(s)

    # Format: "1,466,129.58" (comma as thousands, dot as decimal)
    if ',' in s and '.' in s:
        s = s.replace(',', '')
        return float(s)

    # Format with spaces as thousands separator and dot as decimal
    s = s.replace(' ', '').replace(',', '')

    try:
        return float(s)
    except ValueError:
        return 0.0


def parse_date(val):
    """Parse date from various formats.

    Handles:
    - datetime objects → "YYYY-MM-DD"
    - Strings like "02032\\n 026" → "2026-03-02" (DDMM + YYYY with newline)
    - Strings like "02032026" → "2026-03-02"
    """
    if val is None:
        return None

    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")

    s = str(val).strip()

    # Remove newlines and extra spaces: "02032\n 026" → "02032026"
    s = re.sub(r'\s+', '', s)

    if len(s) == 8 and s.isdigit():
        # Format: DDMMYYYY
        day = s[0:2]
        month = s[2:4]
        year = s[4:8]
        try:
            return f"{year}-{month}-{day}"
        except:
            return s

    return s


def parse_time(val):
    """Parse time value."""
    if val is None:
        return None
    if isinstance(val, dt_time):
        return val.strftime("%H:%M:%S")
    if isinstance(val, datetime):
        return val.strftime("%H:%M:%S")
    return str(val).strip()


def clean_string(val):
    """Clean string value — remove extra whitespace and newlines."""
    if val is None:
        return ""
    s = str(val).strip()
    # Replace newlines and multiple spaces with single space
    s = re.sub(r'\s+', ' ', s)
    return s


def parse_bank_statement(filepath=None):
    """Parse all monthly sheets from Santander bank statement.

    Returns list of dicts with keys:
    month, date, time, branch, description, cargo, abono, balance,
    reference, concept, long_description, comment, dds_category
    """
    filepath = filepath or BANK_STATEMENT_FILE

    if not os.path.exists(filepath):
        print(f"ERROR: File not found: {filepath}")
        return []

    wb = openpyxl.load_workbook(filepath, data_only=True)
    transactions = []

    for sheet_name in wb.sheetnames:
        if sheet_name in BANK_SKIP_SHEETS:
            continue

        ws = wb[sheet_name]

        # Skip sheets with very few rows (metadata)
        if ws.max_row < 3:
            continue

        row_count = 0
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            # Skip empty rows
            if row is None or all(v is None for v in row):
                continue

            # Skip header-like rows (check if first cell looks like account number)
            first_val = row[0] if len(row) > 0 else None
            if first_val is None:
                continue

            # Account number should be numeric
            try:
                float(first_val)
            except (ValueError, TypeError):
                continue

            # Extract columns
            fecha = row[BANK_COLUMNS["Fecha"]] if len(row) > BANK_COLUMNS["Fecha"] else None
            hora = row[BANK_COLUMNS["Hora"]] if len(row) > BANK_COLUMNS["Hora"] else None
            sucursal = row[BANK_COLUMNS["Sucursal"]] if len(row) > BANK_COLUMNS["Sucursal"] else None
            descripcion = row[BANK_COLUMNS["Descripción"]] if len(row) > BANK_COLUMNS["Descripción"] else None
            cargo = row[BANK_COLUMNS["Importe Cargo"]] if len(row) > BANK_COLUMNS["Importe Cargo"] else None
            abono = row[BANK_COLUMNS["Importe Abono"]] if len(row) > BANK_COLUMNS["Importe Abono"] else None
            saldo = row[BANK_COLUMNS["Saldo"]] if len(row) > BANK_COLUMNS["Saldo"] else None
            referencia = row[BANK_COLUMNS["Referencia"]] if len(row) > BANK_COLUMNS["Referencia"] else None
            concepto = row[BANK_COLUMNS["Concepto"]] if len(row) > BANK_COLUMNS["Concepto"] else None
            desc_larga = row[BANK_COLUMNS["Descripción Larga"]] if len(row) > BANK_COLUMNS["Descripción Larga"] else None
            comentario = row[BANK_COLUMNS["Комментарий"]] if len(row) > BANK_COLUMNS["Комментарий"] else None
            dds = row[BANK_COLUMNS["Статья ДДС"]] if len(row) > BANK_COLUMNS["Статья ДДС"] else None

            txn = {
                "month": sheet_name,
                "date": parse_date(fecha),
                "time": parse_time(hora),
                "branch": clean_string(sucursal),
                "description": clean_string(descripcion),
                "cargo": parse_amount(cargo),
                "abono": parse_amount(abono),
                "balance": parse_amount(saldo),
                "reference": clean_string(referencia),
                "concept": clean_string(concepto),
                "long_description": clean_string(desc_larga),
                "comment": clean_string(comentario),
                "dds_category": clean_string(dds),
            }

            transactions.append(txn)
            row_count += 1

    wb.close()
    return transactions


if __name__ == "__main__":
    print("Parsing bank statement...")
    txns = parse_bank_statement()

    print(f"\nTotal transactions: {len(txns)}")

    # Per month summary
    month_counts = {}
    month_cargo = {}
    month_abono = {}
    for t in txns:
        m = t["month"]
        month_counts[m] = month_counts.get(m, 0) + 1
        month_cargo[m] = month_cargo.get(m, 0) + t["cargo"]
        month_abono[m] = month_abono.get(m, 0) + t["abono"]

    print(f"\n{'Month':<20} {'Txns':>6} {'Cargo (MXN)':>15} {'Abono (MXN)':>15}")
    print("-" * 60)
    total_cargo = 0
    total_abono = 0
    for m in month_counts:
        c = month_cargo[m]
        a = month_abono[m]
        total_cargo += c
        total_abono += a
        print(f"{m:<20} {month_counts[m]:>6} {c:>15,.2f} {a:>15,.2f}")

    print("-" * 60)
    print(f"{'TOTAL':<20} {len(txns):>6} {total_cargo:>15,.2f} {total_abono:>15,.2f}")

    # Category coverage
    with_category = sum(1 for t in txns if t["dds_category"])
    print(f"\nWith DDS category: {with_category}/{len(txns)} ({100*with_category/len(txns):.1f}%)")

    # Unique categories
    categories = set(t["dds_category"] for t in txns if t["dds_category"])
    print(f"Unique categories: {len(categories)}")
    for cat in sorted(categories):
        count = sum(1 for t in txns if t["dds_category"] == cat)
        print(f"  {cat}: {count}")
