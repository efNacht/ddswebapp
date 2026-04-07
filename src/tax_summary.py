"""
Tax Summary Generator for FL Cosmetics Mexico.
Generates tax-related summaries from categorized bank transactions:
- НДС (IVA/VAT) breakdown by month
- Tax payment categories (employees, leaders, customs/import)
- Monthly tax obligation summary
"""

import os
import sys
from collections import defaultdict

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from config import OUTPUT_DIR

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
except ImportError:
    print("ERROR: openpyxl not installed.")
    sys.exit(1)

# Tax categories from DDS structure
TAX_CATEGORIES = {
    "НДС": "НДС",
    "Налог на сотрудников": "Налог на сотрудников",
    "Налог на выплаты лидерам": "Налог на выплаты лидерам",
    "Услуги по таможенному оформлению": "Услуги по таможенному оформлению",
}

# VAT rate in Mexico
VAT_RATE = 0.16


def generate_tax_summary(categorized_transactions):
    """Generate monthly tax summary from categorized transactions.

    Returns dict with:
    - monthly_taxes: {month: {tax_category: amount}}
    - monthly_vat_estimate: {month: {revenue: X, vat_on_revenue: Y, vat_paid: Z}}
    - totals: {tax_category: total_amount}
    """
    from dds_generator import MONTH_ORDER

    monthly_taxes = defaultdict(lambda: defaultdict(float))
    monthly_revenue = defaultdict(float)

    for txn in categorized_transactions:
        month = txn["month"]
        cat = txn.get("predicted_category", "")
        cargo = txn.get("cargo", 0) or 0
        abono = txn.get("abono", 0) or 0

        # Tax payments (cargo = expense)
        if cat in TAX_CATEGORIES:
            monthly_taxes[month][cat] += cargo

        # Revenue for VAT estimation (ОП = incoming payments)
        if cat == "ОП":
            monthly_revenue[month] += abono

    # Sort months
    sorted_months = sorted(monthly_taxes.keys(), key=lambda m: MONTH_ORDER.get(m, 99))
    all_months = sorted(
        set(list(monthly_taxes.keys()) + list(monthly_revenue.keys())),
        key=lambda m: MONTH_ORDER.get(m, 99)
    )

    # Compute VAT estimates
    monthly_vat = {}
    for month in all_months:
        revenue = monthly_revenue.get(month, 0)
        # Revenue includes VAT, so base = revenue / 1.16
        revenue_base = revenue / (1 + VAT_RATE)
        vat_on_revenue = revenue - revenue_base
        vat_paid = monthly_taxes[month].get("НДС", 0)

        monthly_vat[month] = {
            "revenue_with_vat": revenue,
            "revenue_base": revenue_base,
            "vat_on_revenue": vat_on_revenue,
            "vat_paid": vat_paid,
            "vat_diff": vat_on_revenue - vat_paid,
        }

    # Totals
    totals = defaultdict(float)
    for month in sorted_months:
        for cat, amount in monthly_taxes[month].items():
            totals[cat] += amount

    return {
        "monthly_taxes": dict(monthly_taxes),
        "monthly_vat": monthly_vat,
        "totals": dict(totals),
        "months": all_months,
    }


def write_tax_excel(tax_data, output_path=None):
    """Write tax summary to Excel."""
    output_path = output_path or os.path.join(OUTPUT_DIR, "Налоговая_сводка.xlsx")

    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(bold=True, size=12)
    section_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    section_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    num_fmt = '#,##0.00'

    months = tax_data["months"]

    # --- Sheet 1: Tax payments by month ---
    ws = wb.active
    ws.title = "Налоговые платежи"

    ws.cell(1, 1, "FL COSMETICS — Налоговые платежи по месяцам").font = header_font
    ws.column_dimensions['A'].width = 40

    r = 3
    # Header row
    cell = ws.cell(r, 1, "Статья")
    cell.fill = header_fill
    cell.font = header_font_white
    cell.border = thin_border
    for j, month in enumerate(months):
        col = 2 + j
        cell = ws.cell(r, col, month)
        cell.fill = header_fill
        cell.font = header_font_white
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 16
    # Total column
    total_col = 2 + len(months)
    cell = ws.cell(r, total_col, "ИТОГО")
    cell.fill = header_fill
    cell.font = header_font_white
    cell.border = thin_border
    ws.column_dimensions[openpyxl.utils.get_column_letter(total_col)].width = 16

    # Data rows
    r += 1
    tax_cats = ["НДС", "Налог на сотрудников", "Налог на выплаты лидерам",
                "Услуги по таможенному оформлению"]

    for cat in tax_cats:
        ws.cell(r, 1, cat).border = thin_border
        row_total = 0
        for j, month in enumerate(months):
            col = 2 + j
            val = tax_data["monthly_taxes"].get(month, {}).get(cat, 0)
            ws.cell(r, col, val).number_format = num_fmt
            ws.cell(r, col).border = thin_border
            row_total += val
        ws.cell(r, total_col, row_total).number_format = num_fmt
        ws.cell(r, total_col).font = Font(bold=True)
        ws.cell(r, total_col).border = thin_border
        r += 1

    # Total row
    ws.cell(r, 1, "ИТОГО налоговые платежи").font = section_font
    ws.cell(r, 1).fill = section_fill
    ws.cell(r, 1).border = thin_border
    grand_total = 0
    for j, month in enumerate(months):
        col = 2 + j
        val = sum(tax_data["monthly_taxes"].get(month, {}).values())
        ws.cell(r, col, val).number_format = num_fmt
        ws.cell(r, col).font = section_font
        ws.cell(r, col).fill = section_fill
        ws.cell(r, col).border = thin_border
        grand_total += val
    ws.cell(r, total_col, grand_total).number_format = num_fmt
    ws.cell(r, total_col).font = Font(bold=True, size=11)
    ws.cell(r, total_col).fill = section_fill
    ws.cell(r, total_col).border = thin_border

    # --- Sheet 2: VAT analysis ---
    ws2 = wb.create_sheet("НДС анализ")
    ws2.column_dimensions['A'].width = 40

    ws2.cell(1, 1, "FL COSMETICS — Анализ НДС (IVA)").font = header_font

    r = 3
    cell = ws2.cell(r, 1, "Показатель")
    cell.fill = header_fill
    cell.font = header_font_white
    cell.border = thin_border
    for j, month in enumerate(months):
        col = 2 + j
        cell = ws2.cell(r, col, month)
        cell.fill = header_fill
        cell.font = header_font_white
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 16
    total_col = 2 + len(months)
    cell = ws2.cell(r, total_col, "ИТОГО")
    cell.fill = header_fill
    cell.font = header_font_white
    cell.border = thin_border
    ws2.column_dimensions[openpyxl.utils.get_column_letter(total_col)].width = 16

    vat_rows = [
        ("Выручка (с НДС)", "revenue_with_vat"),
        ("Выручка (без НДС)", "revenue_base"),
        ("НДС в выручке (16%)", "vat_on_revenue"),
        ("НДС уплачен (факт)", "vat_paid"),
        ("Разница (начислено − уплачено)", "vat_diff"),
    ]

    for label, key in vat_rows:
        r += 1
        ws2.cell(r, 1, label).border = thin_border
        if key == "vat_diff":
            ws2.cell(r, 1).font = section_font
            ws2.cell(r, 1).fill = section_fill
        row_total = 0
        for j, month in enumerate(months):
            col = 2 + j
            val = tax_data["monthly_vat"].get(month, {}).get(key, 0)
            ws2.cell(r, col, val).number_format = num_fmt
            ws2.cell(r, col).border = thin_border
            if key == "vat_diff":
                ws2.cell(r, col).font = section_font
                ws2.cell(r, col).fill = section_fill
            row_total += val
        ws2.cell(r, total_col, row_total).number_format = num_fmt
        ws2.cell(r, total_col).font = Font(bold=True)
        ws2.cell(r, total_col).border = thin_border

    # Notes
    r += 2
    ws2.cell(r, 1, "Примечания:").font = section_font
    r += 1
    ws2.cell(r, 1, "* НДС (IVA) в Мексике = 16%")
    r += 1
    ws2.cell(r, 1, "* Выручка = все поступления по статье ОП из банковской выписки")
    r += 1
    ws2.cell(r, 1, "* НДС уплачен = фактические платежи по статье НДС из выписки")
    r += 1
    ws2.cell(r, 1, "* Разница = сумма к доплате (положительная) или переплата (отрицательная)")

    wb.save(output_path)
    print(f"Tax summary saved to {output_path}")
    return output_path


def print_tax_summary(tax_data):
    """Print tax summary to console."""
    from dds_generator import MONTH_ORDER

    months = tax_data["months"]

    print(f"\n{'='*70}")
    print(f"FL COSMETICS — НАЛОГОВАЯ СВОДКА")
    print(f"{'='*70}")

    # Tax payments
    print(f"\n--- Налоговые платежи (MXN) ---")
    print(f"{'Статья':<35s}", end="")
    for month in months:
        print(f"{month[:8]:>14s}", end="")
    print(f"{'ИТОГО':>14s}")
    print("-" * (35 + 14 * (len(months) + 1)))

    tax_cats = ["НДС", "Налог на сотрудников", "Налог на выплаты лидерам",
                "Услуги по таможенному оформлению"]
    for cat in tax_cats:
        print(f"{cat:<35s}", end="")
        row_total = 0
        for month in months:
            val = tax_data["monthly_taxes"].get(month, {}).get(cat, 0)
            if val:
                print(f"{val:>14,.0f}", end="")
            else:
                print(f"{'':>14s}", end="")
            row_total += val
        print(f"{row_total:>14,.0f}")

    # Total
    print("-" * (35 + 14 * (len(months) + 1)))
    print(f"{'ИТОГО':<35s}", end="")
    grand = 0
    for month in months:
        val = sum(tax_data["monthly_taxes"].get(month, {}).values())
        print(f"{val:>14,.0f}", end="")
        grand += val
    print(f"{grand:>14,.0f}")

    # VAT analysis
    print(f"\n--- НДС (IVA) анализ ---")
    print(f"{'Показатель':<35s}", end="")
    for month in months:
        print(f"{month[:8]:>14s}", end="")
    print(f"{'ИТОГО':>14s}")
    print("-" * (35 + 14 * (len(months) + 1)))

    for label, key in [("НДС в выручке", "vat_on_revenue"),
                       ("НДС уплачен", "vat_paid"),
                       ("Разница", "vat_diff")]:
        print(f"{label:<35s}", end="")
        row_total = 0
        for month in months:
            val = tax_data["monthly_vat"].get(month, {}).get(key, 0)
            if val:
                print(f"{val:>14,.0f}", end="")
            else:
                print(f"{'':>14s}", end="")
            row_total += val
        print(f"{row_total:>14,.0f}")

    print(f"\n{'='*70}")


if __name__ == "__main__":
    from parser import parse_bank_statement
    from categorizer import (
        build_dictionary_from_bank, build_dictionary_from_william,
        merge_dictionaries, categorize_all
    )
    from config import PAYMENTS_FILE

    print("=" * 60)
    print("FL Cosmetics — Tax Summary")
    print("=" * 60)

    # 1. Parse & categorize
    print("\n1. Parsing and categorizing transactions...")
    txns = parse_bank_statement()
    bank_dict = build_dictionary_from_bank(txns)
    william_dict = build_dictionary_from_william(PAYMENTS_FILE)
    merged = merge_dictionaries(bank_dict, william_dict)
    categorized = categorize_all(txns, merged)

    # 2. Generate tax summary
    print("\n2. Generating tax summary...")
    tax_data = generate_tax_summary(categorized)

    # 3. Print
    print_tax_summary(tax_data)

    # 4. Write Excel
    print("\n3. Writing Excel file...")
    path = write_tax_excel(tax_data)
    print(f"\nDone! File: {path}")
