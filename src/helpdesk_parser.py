"""
HELPDESK Data Parsers for FL Cosmetics Mexico.

Parses three HELPDESK export files:
1. Returns/Claims (претензии/возвраты)
2. Write-offs (списания товара)
3. MercadoPago transaction details

Each parser returns a list of dicts + summary stats.
All monetary values are in MXN.
"""

import os
from datetime import datetime
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    raise ImportError("openpyxl is required")


# ─────────────────────────────────────────────────
# 1. Returns / Claims
# ─────────────────────────────────────────────────

def parse_returns(filepath):
    """Parse HELPDESK returns/claims file.

    Expected columns: №, страна, дата возврата, номер заказа,
    артикул, название, количество, сумма возврата, валюта,
    тип претензии, причина, комментарий, статус

    Returns:
        list[dict]: parsed return records
        dict: summary statistics
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Returns file not found: {filepath}")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # Detect header row
    headers = []
    header_row = 1
    for r in range(1, min(5, ws.max_row + 1)):
        row_vals = [str(ws.cell(r, c).value or "").strip().lower() for c in range(1, ws.max_column + 1)]
        if any("возврат" in v or "претензи" in v or "артикул" in v for v in row_vals):
            header_row = r
            headers = row_vals
            break

    if not headers:
        headers = [str(ws.cell(1, c).value or "").strip().lower() for c in range(1, ws.max_column + 1)]

    records = []
    for r in range(header_row + 1, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if not any(row):
            continue

        # Actual structure: col1=rownum, col2=claim#, col3=country, col4=date, ...
        # Columns: №, Номер претензии, Страна, Дата возврата, Номер заказа,
        # Артикул, Название, Количество, Сумма возврата, Валюта,
        # Тип претензии, Причина, Комментарий, Статус
        record = {
            "row_num": row[0],
            "claim_number": row[1] if len(row) > 1 else None,
            "country": row[2] if len(row) > 2 else None,
            "return_date": row[3] if len(row) > 3 else None,
            "order_number": row[4] if len(row) > 4 else None,
            "article": row[5] if len(row) > 5 else None,
            "product_name": row[6] if len(row) > 6 else None,
            "quantity": row[7] if len(row) > 7 else 0,
            "return_amount": row[8] if len(row) > 8 else 0,
            "currency": row[9] if len(row) > 9 else "MXN",
            "claim_type": row[10] if len(row) > 10 else None,
            "reason": row[11] if len(row) > 11 else None,
            "comment": row[12] if len(row) > 12 else None,
            "status": row[13] if len(row) > 13 else None,
        }

        # Normalize amounts
        amt = record["return_amount"]
        record["return_amount"] = float(amt) if isinstance(amt, (int, float)) else 0
        qty = record["quantity"]
        record["quantity"] = int(qty) if isinstance(qty, (int, float)) else 0

        # Normalize date
        if isinstance(record["return_date"], datetime):
            record["return_date_str"] = record["return_date"].strftime("%Y-%m-%d")
            record["return_month"] = _month_name(record["return_date"])
        else:
            record["return_date_str"] = str(record["return_date"] or "")
            record["return_month"] = None

        if record["return_amount"] > 0 or record["quantity"] > 0:
            records.append(record)

    wb.close()

    # Summary
    total_amount = sum(r["return_amount"] for r in records)
    total_qty = sum(r["quantity"] for r in records)
    by_reason = defaultdict(lambda: {"count": 0, "amount": 0, "qty": 0})
    by_month = defaultdict(lambda: {"count": 0, "amount": 0, "qty": 0})
    by_status = defaultdict(int)

    for r in records:
        reason = str(r.get("reason") or r.get("claim_type") or "Не указана")
        by_reason[reason]["count"] += 1
        by_reason[reason]["amount"] += r["return_amount"]
        by_reason[reason]["qty"] += r["quantity"]

        month = r.get("return_month") or "Unknown"
        by_month[month]["count"] += 1
        by_month[month]["amount"] += r["return_amount"]
        by_month[month]["qty"] += r["quantity"]

        status = str(r.get("status") or "Unknown")
        by_status[status] += 1

    summary = {
        "total_records": len(records),
        "total_amount": total_amount,
        "total_quantity": total_qty,
        "by_reason": dict(by_reason),
        "by_month": dict(by_month),
        "by_status": dict(by_status),
    }

    return records, summary


# ─────────────────────────────────────────────────
# 2. Write-offs
# ─────────────────────────────────────────────────

def parse_writeoffs(filepath):
    """Parse HELPDESK write-offs file.

    Expected columns: №, Склад, Статья расходов, Дата, Артикул,
    Название, Количество, с/с единицы, Сумма по с/с, Валюта, Причина списания

    Returns:
        list[dict]: parsed write-off records
        dict: summary with breakdowns
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Write-offs file not found: {filepath}")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # Find header row
    header_row = 1
    for r in range(1, min(5, ws.max_row + 1)):
        row_vals = [str(ws.cell(r, c).value or "").strip().lower() for c in range(1, ws.max_column + 1)]
        if any("склад" in v or "артикул" in v or "списан" in v for v in row_vals):
            header_row = r
            break

    records = []
    for r in range(header_row + 1, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if not any(row):
            continue

        record = {
            "row_num": row[0],
            "warehouse": row[1] if len(row) > 1 else None,
            "expense_category": row[2] if len(row) > 2 else None,
            "date": row[3] if len(row) > 3 else None,
            "article": row[4] if len(row) > 4 else None,
            "product_name": row[5] if len(row) > 5 else None,
            "quantity": row[6] if len(row) > 6 else 0,
            "unit_cost": row[7] if len(row) > 7 else 0,
            "total_cost": row[8] if len(row) > 8 else 0,
            "currency": row[9] if len(row) > 9 else "MXN",
            "writeoff_reason": row[10] if len(row) > 10 else None,
        }

        # Normalize
        for field in ["quantity", "unit_cost", "total_cost"]:
            v = record[field]
            record[field] = float(v) if isinstance(v, (int, float)) else 0

        if isinstance(record["date"], datetime):
            record["date_str"] = record["date"].strftime("%Y-%m-%d")
            record["month"] = _month_name(record["date"])
        else:
            record["date_str"] = str(record["date"] or "")
            record["month"] = None

        if record["total_cost"] > 0 or record["quantity"] > 0:
            records.append(record)

    wb.close()

    # Summary
    total_cost = sum(r["total_cost"] for r in records)
    total_qty = sum(r["quantity"] for r in records)

    by_reason = defaultdict(lambda: {"count": 0, "cost": 0, "qty": 0})
    by_category = defaultdict(lambda: {"count": 0, "cost": 0, "qty": 0})
    by_month = defaultdict(lambda: {"count": 0, "cost": 0, "qty": 0})

    for r in records:
        reason = str(r.get("writeoff_reason") or "Не указана")
        by_reason[reason]["count"] += 1
        by_reason[reason]["cost"] += r["total_cost"]
        by_reason[reason]["qty"] += r["quantity"]

        cat = str(r.get("expense_category") or "Не указана")
        by_category[cat]["count"] += 1
        by_category[cat]["cost"] += r["total_cost"]
        by_category[cat]["qty"] += r["quantity"]

        month = r.get("month") or "Unknown"
        by_month[month]["count"] += 1
        by_month[month]["cost"] += r["total_cost"]
        by_month[month]["qty"] += r["quantity"]

    summary = {
        "total_records": len(records),
        "total_cost": total_cost,
        "total_quantity": total_qty,
        "by_reason": dict(by_reason),
        "by_category": dict(by_category),
        "by_month": dict(by_month),
    }

    return records, summary


# ─────────────────────────────────────────────────
# 3. MercadoPago Details
# ─────────────────────────────────────────────────

def parse_mercadopago(filepath):
    """Parse HELPDESK MercadoPago transaction details.

    Expected columns: Провайдер, Номер заказа, Дата, Сумма поступлений,
    Сумма отмен, Кол-во транзакций, Комиссия, Валюта

    Returns:
        list[dict]: parsed MP records
        dict: summary with monthly breakdown
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"MercadoPago file not found: {filepath}")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # Find header row
    header_row = 1
    for r in range(1, min(5, ws.max_row + 1)):
        row_vals = [str(ws.cell(r, c).value or "").strip().lower() for c in range(1, ws.max_column + 1)]
        if any("провайдер" in v or "mercado" in v or "поступлен" in v for v in row_vals):
            header_row = r
            break

    records = []
    for r in range(header_row + 1, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if not any(row):
            continue

        # Actual structure: col1=rownum, col2=provider, col3=order#, col4=date,
        # col5=revenue, col6=cancellations, col7=txn_count, col8=commission, col9=currency
        record = {
            "row_num": row[0],
            "provider": row[1] if len(row) > 1 else "MercadoPago",
            "order_number": row[2] if len(row) > 2 else None,
            "date": row[3] if len(row) > 3 else None,
            "revenue": row[4] if len(row) > 4 else 0,
            "cancellations": row[5] if len(row) > 5 else 0,
            "txn_count": row[6] if len(row) > 6 else 0,
            "commission": row[7] if len(row) > 7 else 0,
            "currency": row[8] if len(row) > 8 else "MXN",
        }

        # Normalize
        for field in ["revenue", "cancellations", "commission"]:
            v = record[field]
            record[field] = float(v) if isinstance(v, (int, float)) else 0
        v = record["txn_count"]
        record["txn_count"] = int(v) if isinstance(v, (int, float)) else 0

        if isinstance(record["date"], datetime):
            record["date_str"] = record["date"].strftime("%Y-%m-%d")
            record["month"] = _month_name(record["date"])
        else:
            record["date_str"] = str(record["date"] or "")
            record["month"] = None

        # Net = revenue - cancellations - commission
        record["net_revenue"] = record["revenue"] - record["cancellations"] - record["commission"]

        if record["revenue"] > 0 or record["cancellations"] > 0:
            records.append(record)

    wb.close()

    # Summary
    total_revenue = sum(r["revenue"] for r in records)
    total_cancel = sum(r["cancellations"] for r in records)
    total_commission = sum(r["commission"] for r in records)
    total_net = sum(r["net_revenue"] for r in records)
    total_txns = sum(r["txn_count"] for r in records)

    by_month = defaultdict(lambda: {"revenue": 0, "cancellations": 0, "commission": 0, "net": 0, "count": 0, "txns": 0})
    for r in records:
        month = r.get("month") or "Unknown"
        by_month[month]["revenue"] += r["revenue"]
        by_month[month]["cancellations"] += r["cancellations"]
        by_month[month]["commission"] += r["commission"]
        by_month[month]["net"] += r["net_revenue"]
        by_month[month]["count"] += 1
        by_month[month]["txns"] += r["txn_count"]

    summary = {
        "total_records": len(records),
        "total_revenue": total_revenue,
        "total_cancellations": total_cancel,
        "total_commission": total_commission,
        "total_net_revenue": total_net,
        "total_transactions": total_txns,
        "commission_rate": (total_commission / total_revenue * 100) if total_revenue > 0 else 0,
        "by_month": dict(by_month),
    }

    return records, summary


# ─────────────────────────────────────────────────
# Excel report writer
# ─────────────────────────────────────────────────

def write_helpdesk_report(returns_summary, writeoffs_summary, mp_summary, output_path):
    """Write combined HELPDESK data report to Excel.

    Creates sheets:
    1. Сводка — overview of all 3 data sources
    2. Возвраты — returns by month/reason
    3. Списания — write-offs by month/category
    4. MercadoPago — MP by month (gross/net/commission)
    """
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # Styles
    h1 = Font(bold=True, size=13)
    h2 = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    red_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    num_fmt = '#,##0.00'
    pct_fmt = '0.0%'

    def _header_row(ws, row, headers, widths=None):
        for j, hdr in enumerate(headers, 1):
            c = ws.cell(row, j, hdr)
            c.fill = header_fill
            c.font = header_font
            c.border = thin
            c.alignment = Alignment(horizontal='center')
            if widths and j <= len(widths):
                ws.column_dimensions[get_column_letter(j)].width = widths[j - 1]

    def _data_cell(ws, row, col, val, fmt=None, bold=False, fill=None):
        c = ws.cell(row, col, val)
        c.border = thin
        if fmt:
            c.number_format = fmt
        if bold:
            c.font = Font(bold=True)
        if fill:
            c.fill = fill
        return c

    # ─── Sheet 1: Сводка ───
    ws = wb.active
    ws.title = "Сводка"
    ws.cell(1, 1, "FL COSMETICS — Данные партнёра (HELPDESK)").font = h1

    r = 3
    ws.cell(r, 1, "Источник данных").font = h2
    ws.cell(r, 2, "Записей").font = h2
    ws.cell(r, 3, "Сумма (MXN)").font = h2
    ws.cell(r, 4, "Кол-во ед.").font = h2
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 14

    r = 4
    if returns_summary:
        _data_cell(ws, r, 1, "Возвраты / Претензии")
        _data_cell(ws, r, 2, returns_summary["total_records"])
        _data_cell(ws, r, 3, returns_summary["total_amount"], num_fmt)
        _data_cell(ws, r, 4, returns_summary["total_quantity"])
        r += 1

    if writeoffs_summary:
        _data_cell(ws, r, 1, "Списания товара")
        _data_cell(ws, r, 2, writeoffs_summary["total_records"])
        _data_cell(ws, r, 3, writeoffs_summary["total_cost"], num_fmt)
        _data_cell(ws, r, 4, writeoffs_summary["total_quantity"])
        r += 1

    if mp_summary:
        _data_cell(ws, r, 1, "MercadoPago — выручка (gross)")
        _data_cell(ws, r, 2, mp_summary["total_records"])
        _data_cell(ws, r, 3, mp_summary["total_revenue"], num_fmt)
        _data_cell(ws, r, 4, mp_summary["total_transactions"])
        r += 1
        _data_cell(ws, r, 1, "MercadoPago — комиссия")
        _data_cell(ws, r, 2, "")
        _data_cell(ws, r, 3, mp_summary["total_commission"], num_fmt, fill=red_fill)
        _data_cell(ws, r, 4, f'{mp_summary["commission_rate"]:.1f}%')
        r += 1
        _data_cell(ws, r, 1, "MercadoPago — отмены")
        _data_cell(ws, r, 2, "")
        _data_cell(ws, r, 3, mp_summary["total_cancellations"], num_fmt, fill=yellow_fill)
        r += 1
        _data_cell(ws, r, 1, "MercadoPago — нетто")
        _data_cell(ws, r, 2, "")
        _data_cell(ws, r, 3, mp_summary["total_net_revenue"], num_fmt, bold=True, fill=green_fill)

    # ─── Sheet 2: Возвраты ───
    if returns_summary and returns_summary["total_records"] > 0:
        ws2 = wb.create_sheet("Возвраты")
        ws2.cell(1, 1, "Возвраты / Претензии по месяцам").font = h1

        r = 3
        _header_row(ws2, r, ["Месяц", "Записей", "Сумма (MXN)", "Количество ед."], [22, 12, 18, 16])
        r += 1
        months_sorted = _sort_months(returns_summary["by_month"].keys())
        for month in months_sorted:
            d = returns_summary["by_month"][month]
            _data_cell(ws2, r, 1, month)
            _data_cell(ws2, r, 2, d["count"])
            _data_cell(ws2, r, 3, d["amount"], num_fmt)
            _data_cell(ws2, r, 4, d["qty"])
            r += 1
        _data_cell(ws2, r, 1, "ИТОГО", bold=True)
        _data_cell(ws2, r, 2, returns_summary["total_records"], bold=True)
        _data_cell(ws2, r, 3, returns_summary["total_amount"], num_fmt, bold=True)
        _data_cell(ws2, r, 4, returns_summary["total_quantity"], bold=True)

        # By reason
        r += 2
        ws2.cell(r, 1, "По причинам").font = h2
        r += 1
        _header_row(ws2, r, ["Причина", "Записей", "Сумма (MXN)", "Количество ед."], [40, 12, 18, 16])
        r += 1
        for reason, d in sorted(returns_summary["by_reason"].items(), key=lambda x: -x[1]["amount"]):
            _data_cell(ws2, r, 1, reason)
            _data_cell(ws2, r, 2, d["count"])
            _data_cell(ws2, r, 3, d["amount"], num_fmt)
            _data_cell(ws2, r, 4, d["qty"])
            r += 1

    # ─── Sheet 3: Списания ───
    if writeoffs_summary and writeoffs_summary["total_records"] > 0:
        ws3 = wb.create_sheet("Списания")
        ws3.cell(1, 1, "Списания товара по месяцам").font = h1

        r = 3
        _header_row(ws3, r, ["Месяц", "Записей", "С/с (MXN)", "Количество ед."], [22, 12, 18, 16])
        r += 1
        months_sorted = _sort_months(writeoffs_summary["by_month"].keys())
        for month in months_sorted:
            d = writeoffs_summary["by_month"][month]
            _data_cell(ws3, r, 1, month)
            _data_cell(ws3, r, 2, d["count"])
            _data_cell(ws3, r, 3, d["cost"], num_fmt)
            _data_cell(ws3, r, 4, d["qty"])
            r += 1
        _data_cell(ws3, r, 1, "ИТОГО", bold=True)
        _data_cell(ws3, r, 2, writeoffs_summary["total_records"], bold=True)
        _data_cell(ws3, r, 3, writeoffs_summary["total_cost"], num_fmt, bold=True)
        _data_cell(ws3, r, 4, writeoffs_summary["total_quantity"], bold=True)

        # By category
        r += 2
        ws3.cell(r, 1, "По статьям расходов").font = h2
        r += 1
        _header_row(ws3, r, ["Статья", "Записей", "С/с (MXN)", "Количество ед."], [45, 12, 18, 16])
        r += 1
        for cat, d in sorted(writeoffs_summary["by_category"].items(), key=lambda x: -x[1]["cost"]):
            _data_cell(ws3, r, 1, cat)
            _data_cell(ws3, r, 2, d["count"])
            _data_cell(ws3, r, 3, d["cost"], num_fmt)
            _data_cell(ws3, r, 4, d["qty"])
            r += 1

        # By reason
        r += 2
        ws3.cell(r, 1, "По причинам списания").font = h2
        r += 1
        _header_row(ws3, r, ["Причина", "Записей", "С/с (MXN)", "Количество ед."], [45, 12, 18, 16])
        r += 1
        for reason, d in sorted(writeoffs_summary["by_reason"].items(), key=lambda x: -x[1]["cost"]):
            _data_cell(ws3, r, 1, reason)
            _data_cell(ws3, r, 2, d["count"])
            _data_cell(ws3, r, 3, d["cost"], num_fmt)
            _data_cell(ws3, r, 4, d["qty"])
            r += 1

    # ─── Sheet 4: MercadoPago ───
    if mp_summary and mp_summary["total_records"] > 0:
        ws4 = wb.create_sheet("MercadoPago")
        ws4.cell(1, 1, "MercadoPago — помесячная аналитика").font = h1

        r = 3
        _header_row(ws4, r,
                    ["Месяц", "Заказов", "Транзакций", "Выручка gross", "Отмены", "Комиссия", "Нетто", "Комиссия %"],
                    [22, 12, 14, 18, 16, 16, 18, 14])
        r += 1
        months_sorted = _sort_months(mp_summary["by_month"].keys())
        for month in months_sorted:
            d = mp_summary["by_month"][month]
            comm_pct = (d["commission"] / d["revenue"] * 100) if d["revenue"] > 0 else 0
            _data_cell(ws4, r, 1, month)
            _data_cell(ws4, r, 2, d["count"])
            _data_cell(ws4, r, 3, d["txns"])
            _data_cell(ws4, r, 4, d["revenue"], num_fmt)
            _data_cell(ws4, r, 5, d["cancellations"], num_fmt, fill=yellow_fill if d["cancellations"] > 0 else None)
            _data_cell(ws4, r, 6, d["commission"], num_fmt, fill=red_fill)
            _data_cell(ws4, r, 7, d["net"], num_fmt, bold=True, fill=green_fill)
            _data_cell(ws4, r, 8, comm_pct / 100, pct_fmt)
            r += 1

        _data_cell(ws4, r, 1, "ИТОГО", bold=True)
        _data_cell(ws4, r, 2, mp_summary["total_records"], bold=True)
        _data_cell(ws4, r, 3, mp_summary["total_transactions"], bold=True)
        _data_cell(ws4, r, 4, mp_summary["total_revenue"], num_fmt, bold=True)
        _data_cell(ws4, r, 5, mp_summary["total_cancellations"], num_fmt, bold=True)
        _data_cell(ws4, r, 6, mp_summary["total_commission"], num_fmt, bold=True)
        _data_cell(ws4, r, 7, mp_summary["total_net_revenue"], num_fmt, bold=True)
        overall_pct = mp_summary["commission_rate"] / 100
        _data_cell(ws4, r, 8, overall_pct, pct_fmt, bold=True)

    wb.save(output_path)
    print(f"[HELPDESK] Report saved: {output_path}")
    return output_path


# ─────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────

MONTH_NAMES_RU = {
    1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель",
    5: "Май", 6: "Июнь", 7: "Июль", 8: "Август",
    9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь",
}

MONTH_ORDER = {}
for y in range(2024, 2028):
    for m in range(1, 13):
        MONTH_ORDER[f"{MONTH_NAMES_RU[m]} {y}"] = y * 100 + m


def _month_name(dt):
    """Convert datetime to 'Месяц YYYY' format."""
    if not isinstance(dt, datetime):
        return None
    return f"{MONTH_NAMES_RU.get(dt.month, '?')} {dt.year}"


def _sort_months(months):
    """Sort month names chronologically."""
    return sorted(months, key=lambda m: MONTH_ORDER.get(m, 99999))
