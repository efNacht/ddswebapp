"""
Template Filler — writes fact data into Emil's original DDS and PL templates.

Copies the template file, then fills in the "факт" columns with computed values.
Preserves all existing formatting, formulas, and plan data.
"""

import os
import sys
import shutil
from collections import defaultdict

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import config as _cfg

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed.")
    sys.exit(1)

# ─────────────────────────────────────────────────────────────
# DDS Template mapping
# ─────────────────────────────────────────────────────────────

# Month → fact column in DDS template (sheet "ДДС 2025-2026_нац валюта")
# Layout: each month has [план, факт, +/-] except quarter/half-year summary cols
# Row 13 shows: c9=план, c10=факт, c12=план, c13=факт, ...
DDS_MONTH_FACT_COL = {
    "Январь 2025": 10,
    "Февраль 2025": 13,
    "Март 2025": 16,
    # 1кв c18(план)/c19(факт) — auto-sum, skip
    "Апрель 2025": 22,
    "Май 2025": 25,
    "Июнь 2025": 28,
    # 2кв c30/c31 — auto-sum
    # 1п c33/c34 — auto-sum
    "Июль 2025": 37,
    "Август 2025": 40,
    "Сентябрь 2025": 43,
    # 3кв c45/c46
    "Октябрь 2025": 49,
    "Ноябрь 2025": 52,
    "Декабрь 2025": 55,
    # 4кв c57, 2п c60, год c63
    "Январь 2026": 66,
    "Февраль 2026": 69,
    "Март 2026": 72,
}

# DDS template row mapping: category_key → row number in template
# Based on the template structure exploration
DDS_ROW_MAP = {
    # r15: объем продаж
    "total_abono": 15,
    # r18: расчетный счет (opening balance)
    "opening_balance": 18,
    # r26: выручка ДП
    "ОП": 26,
    # r28: выплата ОС по МП (контракты)
    "ОС лидерам": 28,
    # r29: квалификационный бонус
    "_кв_бонус": 29,
    # r32: приход прочий
    "Взнос в УК": 32,
    # r56: услуги по таможенному оформлению, транспортные расходы (импорт)
    "Услуги по таможенному оформлению": 56,
    # r63: заработная плата сотрудников (office)
    "ЗП офис": 63,
    # r78: стимулирование продаж
    "Стимулирование продаж": 78,
    # r81: услуги банка (РКО)
    "Услуги банка": 81,
    # r82: услуги связи, почта, интернет
    "Связь": 82,
    # r83: Бухгалтерские услуги
    "Бухгалтерские услуги": 83,
    # r88: аренда склада
    "Аренда склада": 88,
    # r89: коммунальные расходы
    "Коммунальные платежи": 89,
    # r90: заработная плата (склад)
    "ЗП склад": 90,
    # r96: прочие расходы
    "Прочие расходы": 96,
    # r97: расходные материалы
    "Расходные материалы": 97,
    # r102: Транспортные в регион
    "Транспортные в регион": 102,
    # r105: хозяйственные расходы
    "Хозяйственные расходы": 105,
    # r116: реклама: производство каталог
    "Реклама каталог": 116,
    # r119: НДС
    "НДС": 119,
    # r120: Налоги за сотрудников
    "Налог на сотрудников": 120,
    # r121: Налоги на выплаты лидерам
    "Налог на выплаты лидерам": 121,
    # r159: расчетный счет (closing balance)
    "closing_balance": 159,
}

# Categories where negative sign = expense (cargo column) in bank → write as positive in template
DDS_EXPENSE_CATEGORIES = {
    "ОС лидерам", "Услуги по таможенному оформлению",
    "ЗП офис", "Стимулирование продаж", "Услуги банка", "Связь",
    "Бухгалтерские услуги", "Аренда склада", "Коммунальные платежи",
    "ЗП склад", "Прочие расходы", "Расходные материалы",
    "Транспортные в регион", "Хозяйственные расходы",
    "Реклама каталог", "НДС", "Налог на сотрудников",
    "Налог на выплаты лидерам",
}

# ─────────────────────────────────────────────────────────────
# PL Template mapping
# ─────────────────────────────────────────────────────────────

# Month → fact column in PL template
# PL sheets: "2025" and "2026"
# Row 7 shows: c11=факт, c13=факт, c15=факт, c19=факт, c21=факт, c23=факт
# Months go in odd columns starting from 11: 11,13,15 (Q1), 19,21,23 (Q2), ...
PL_2025_MONTH_COL = {
    "Январь 2025": 11,     # not in bank (no data)
    "Февраль 2025": 11,    # actually Jan=11, Feb=13
    "Март 2025": 15,
    "Апрель 2025": 19,
    "Май 2025": 21,
    "Июнь 2025": 23,
    "Июль 2025": 29,       # need to check exact columns
    "Август 2025": 31,
    "Сентябрь 2025": 33,
    "Октябрь 2025": 39,
    "Ноябрь 2025": 41,
    "Декабрь 2025": 43,
}

PL_2026_MONTH_COL = {
    "Январь 2026": 11,
    "Февраль 2026": 13,
    "Март 2026": 15,
}

# PL row mapping: what to fill and from which DDS category
# Only filling rows that come from bank data
PL_ROW_MAP = {
    # r2/c9: Объем продаж → total_abono (same as ОП)
    # r8: ОП без НДС = revenue / 1.16
    # r76: аренда склада
    "Аренда склада": 76,
    # r77: расходные материалы
    "Расходные материалы": 77,
    # r80: интернет и связь
    "Связь": 80,
    # r81: хозяйственные расходы
    "Хозяйственные расходы": 81,
    # r82: зарплата и ФОТ склад
    "ЗП склад": 82,
    # r90: транспортные (в регионы)
    "Транспортные в регион": 90,
    # r109: заработная плата (office)
    "ЗП офис": 109,
    # r125: аудит, бух.учет
    "Бухгалтерские услуги": 125,
    # r126: коммунальные расходы
    "Коммунальные платежи": 126,
    # r131: стимулирование продаж
    "Стимулирование продаж": 131,
    # r134: услуги по терм.оплате  (эквайринг)
    # r135: производство полиграфической продукции
    "Реклама каталог": 135,
    # r140: налоги фот
    "Налог на сотрудников": 140,
}


def fill_dds_template(dds_data, output_path=None, template_path=None):
    """Fill the original DDS template with fact data.

    Copies the template file and fills fact columns.
    Returns path to the filled file.
    """
    output_path = output_path or os.path.join(_cfg.OUTPUT_DIR, "ДДС_план_факт.xlsx")
    template_path = template_path or _cfg.DDS_TEMPLATE_FILE

    print(f"[DDS_FILL] template_path={template_path} exists={os.path.exists(template_path) if template_path else False}", flush=True)

    if not template_path or not os.path.exists(template_path):
        raise FileNotFoundError(f"DDS template not found: {template_path}")

    # Copy template
    shutil.copy2(template_path, output_path)

    # Open with openpyxl (keep formatting)
    wb = openpyxl.load_workbook(output_path)

    # Find the DDS sheet — try exact name first, then fuzzy match
    ws = None
    for name in wb.sheetnames:
        if name == "ДДС 2025-2026_нац валюта" or "ДДС" in name or "DDS" in name.upper():
            ws = wb[name]
            print(f"[DDS_FILL] Using sheet: '{name}'", flush=True)
            break
    if ws is None:
        raise KeyError(f"DDS sheet not found. Available sheets: {wb.sheetnames}")

    filled_count = 0

    for month, col in DDS_MONTH_FACT_COL.items():
        if month not in dds_data:
            continue

        month_data = dds_data[month]

        for cat_key, row in DDS_ROW_MAP.items():
            value = month_data.get(cat_key, 0)
            if value and value != 0:
                ws.cell(row, col, value)
                filled_count += 1

    wb.save(output_path)
    wb.close()
    print(f"DDS template filled: {filled_count} cells → {output_path}")
    return output_path


def fill_pl_template(categorized_transactions, output_path=None, template_path=None):
    """Fill the original PL template with fact data from bank.

    Only fills expense rows that come from bank categorization.
    Revenue/selfcost rows need Sales Reports (filled separately).
    """
    from dds_generator import MONTH_ORDER, aggregate_by_month

    output_path = output_path or os.path.join(_cfg.OUTPUT_DIR, "PL_план_факт.xlsx")
    template_path = template_path or _cfg.PL_TEMPLATE_FILE

    print(f"[PL_FILL] template_path={template_path} exists={os.path.exists(template_path) if template_path else False}", flush=True)

    if not template_path or not os.path.exists(template_path):
        raise FileNotFoundError(f"PL template not found: {template_path}")

    # Copy template
    shutil.copy2(template_path, output_path)

    wb = openpyxl.load_workbook(output_path)

    monthly, balances = aggregate_by_month(categorized_transactions)

    filled_count = 0

    # Fill sheet "2025"
    if "2025" in wb.sheetnames:
        ws = wb["2025"]
        # Verify exact month columns by reading row 1 dates
        # Row 1: c11=2025-01-01, c13=2025-02-01, c15=2025-03-01, c19=2025-04-01, ...
        # Map month names to columns by reading dates
        month_cols_2025 = _detect_month_columns(ws, row=1)

        for month, col in month_cols_2025.items():
            if month not in monthly:
                continue
            cats = monthly[month]

            # Fill expense rows
            for cat_key, pl_row in PL_ROW_MAP.items():
                cargo = cats.get(cat_key, {}).get("cargo", 0)
                if cargo and cargo != 0:
                    ws.cell(pl_row, col, cargo)
                    filled_count += 1

            # Revenue rows
            op_abono = cats.get("ОП", {}).get("abono", 0)
            if op_abono:
                # r2: Объем продаж
                ws.cell(2, col, op_abono)
                filled_count += 1

    # Fill sheet "2026"
    if "2026" in wb.sheetnames:
        ws = wb["2026"]
        month_cols_2026 = _detect_month_columns(ws, row=1)

        for month, col in month_cols_2026.items():
            if month not in monthly:
                continue
            cats = monthly[month]

            for cat_key, pl_row in PL_ROW_MAP.items():
                cargo = cats.get(cat_key, {}).get("cargo", 0)
                if cargo and cargo != 0:
                    ws.cell(pl_row, col, cargo)
                    filled_count += 1

            op_abono = cats.get("ОП", {}).get("abono", 0)
            if op_abono:
                ws.cell(2, col, op_abono)
                filled_count += 1

    wb.save(output_path)
    wb.close()
    print(f"PL template filled: {filled_count} cells → {output_path}")
    return output_path


def _detect_month_columns(ws, row=1):
    """Auto-detect month → column mapping from PL template header row.

    Row 1 contains dates like 2025-01-01, 2025-02-01, etc.
    Returns dict: {"Февраль 2025": 13, "Март 2025": 15, ...}
    """
    from datetime import datetime

    RUSSIAN_MONTHS = {
        1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель",
        5: "Май", 6: "Июнь", 7: "Июль", 8: "Август",
        9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь",
    }

    result = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row, c).value
        if isinstance(val, datetime):
            month_name = f"{RUSSIAN_MONTHS[val.month]} {val.year}"
            result[month_name] = c
    return result


if __name__ == "__main__":
    from parser import parse_bank_statement
    from categorizer import (
        build_dictionary_from_bank, build_dictionary_from_william,
        merge_dictionaries, categorize_all
    )
    from config import PAYMENTS_FILE
    from dds_generator import generate_dds_data

    print("=" * 60)
    print("FL Cosmetics — Template Filler")
    print("=" * 60)

    # 1. Parse & categorize
    print("\n1. Parsing and categorizing transactions...")
    txns = parse_bank_statement()
    bank_dict = build_dictionary_from_bank(txns)
    william_dict = build_dictionary_from_william(PAYMENTS_FILE)
    merged = merge_dictionaries(bank_dict, william_dict)
    categorized = categorize_all(txns, merged)

    # 2. Fill DDS template
    print("\n2. Filling DDS template...")
    dds_data = generate_dds_data(categorized)
    dds_path = fill_dds_template(dds_data)

    # 3. Fill PL template
    print("\n3. Filling PL template...")
    pl_path = fill_pl_template(categorized)

    print(f"\nDone!")
    print(f"  DDS: {dds_path}")
    print(f"  PL:  {pl_path}")
