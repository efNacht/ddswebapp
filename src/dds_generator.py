"""
ДДС (Cash Flow Statement) Generator for FL Cosmetics.
Maps categorized bank transactions to DDS template structure.

DDS template structure (from "ДДС Плановый Мексика 2025"):
- Объем продаж (sales volume)
- Остаток ДС начало (opening balance)
- Денежный поток от операционной деятельности:
  - Приток: выручка ДП, выплата КВ, выплата ОС, кв.бонус
  - Отток: закупка товаров, коммерческие расходы (офис + логистика + реклама), налоги
- Денежный поток от инвестиционной деятельности
- Денежный поток от финансовой деятельности
- Остаток ДС конец (closing balance)
"""

import json
import os
import sys
from collections import defaultdict
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from config import OUTPUT_DIR, DDS_TEMPLATE_FILE

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip3 install openpyxl")
    sys.exit(1)


# --- Category → DDS line mapping ---
# Maps our normalized categories to DDS template rows

# DDS structure with hierarchy
DDS_STRUCTURE = [
    # (level, label, category_key, is_income)
    # level: indentation depth (1=section, 2=subsection, 3=line item)
    # category_key: key used to sum transactions, None for subtotal rows
    # is_income: True if we use abono, False if we use cargo

    (1, "Объем продаж", None, None),
    (2, "объем продаж", "total_abono", True),

    (1, "Остаток ДС начало периода", None, None),
    (2, "расчетный счет", "opening_balance", None),

    (1, "Денежный поток", None, None),
    (2, "Операционная деятельность", None, None),
    (3, "Приток ДС", None, None),
    (4, "Выручка от реализации", None, None),
    (5, "выручка ДП", "ОП", True),
    (5, "выплата ОС по контрактам", "ОС лидерам", False),
    (5, "квалификационный бонус", "_кв_бонус", False),

    (4, "Прочие поступления", None, None),
    (5, "взнос в УК", "Взнос в УК", True),
    (5, "возврат средств", "Возврат средств", True),

    (3, "Отток ДС", None, None),
    (4, "Закупка товаров, НДС импорт", None, None),
    (5, "закупка товара (контейнер)", "Закупка товара (контейнер)", False),
    (5, "услуги по таможенному оформлению", "Услуги по таможенному оформлению", False),

    (4, "Коммерческие расходы офис", None, None),
    (5, "заработная плата сотрудников (офис)", "ЗП офис", False),
    (5, "стимулирование продаж", "Стимулирование продаж", False),
    (5, "услуги банка (РКО)", "Услуги банка", False),
    (5, "услуги связи, почта, интернет", "Связь", False),
    (5, "бухгалтерские услуги", "Бухгалтерские услуги", False),
    (5, "прочие расходы", "Прочие расходы", False),

    (4, "Коммерческие расходы логистика", None, None),
    (5, "аренда склада", "Аренда склада", False),
    (5, "коммунальные расходы", "Коммунальные платежи", False),
    (5, "заработная плата склад", "ЗП склад", False),
    (5, "расходные материалы", "Расходные материалы", False),
    (5, "транспортные в регион", "Транспортные в регион", False),
    (5, "хозяйственные расходы", "Хозяйственные расходы", False),
    (5, "упаковка", "Упаковка", False),

    (4, "Реклама и маркетинг", None, None),
    (5, "реклама каталог", "Реклама каталог", False),

    (4, "Выплаты в бюджет", None, None),
    (5, "НДС", "НДС", False),
    (5, "Налоги за сотрудников", "Налог на сотрудников", False),
    (5, "Налоги на выплаты лидерам", "Налог на выплаты лидерам", False),

    (1, "Остаток ДС конец периода", None, None),
    (2, "расчетный счет", "closing_balance", None),
]

# Month name → order mapping (for sorting)
MONTH_ORDER = {
    "Январь 2025": 1, "Февраль 2025": 2, "Март 2025": 3,
    "Апрель 2025": 4, "Май 2025": 5, "Июнь 2025": 6,
    "Июль 2025": 7, "Август 2025": 8, "Сентябрь 2025": 9,
    "Октябрь 2025": 10, "Ноябрь 2025": 11, "Декабрь 2025": 12,
    "Январь 2026": 13, "Февраль 2026": 14, "Март 2026": 15,
}


def aggregate_by_month(categorized_transactions):
    """Aggregate categorized transactions by month and category.

    Returns dict: {month_name: {category: {"cargo": sum, "abono": sum, "count": N}}}
    Also computes opening/closing balances per month.
    """
    from categorizer import normalize_category

    monthly = defaultdict(lambda: defaultdict(lambda: {"cargo": 0.0, "abono": 0.0, "count": 0}))
    monthly_balances = defaultdict(list)

    for txn in categorized_transactions:
        month = txn["month"]
        cat = txn.get("predicted_category", "UNKNOWN")
        if not cat:
            cat = "UNKNOWN"

        monthly[month][cat]["cargo"] += txn.get("cargo", 0) or 0
        monthly[month][cat]["abono"] += txn.get("abono", 0) or 0
        monthly[month][cat]["count"] += 1

        # Track balance for opening/closing
        balance = txn.get("balance", 0) or 0
        if balance > 0:
            monthly_balances[month].append(balance)

        # Also track total income/expense
        monthly[month]["_total"]["cargo"] += txn.get("cargo", 0) or 0
        monthly[month]["_total"]["abono"] += txn.get("abono", 0) or 0
        monthly[month]["_total"]["count"] += 1

    # Compute opening/closing balances
    balances = {}
    for month in monthly:
        if monthly_balances[month]:
            # Last balance in month = closing
            balances[month] = {
                "closing": monthly_balances[month][-1],
                # Opening = first balance + first cargo - first abono (approximate)
                "opening": monthly_balances[month][0],
            }
        else:
            balances[month] = {"opening": 0, "closing": 0}

    return monthly, balances


def generate_dds_data(categorized_transactions):
    """Generate DDS fact data from categorized transactions.

    Returns dict: {month: {dds_line_key: amount}}
    """
    monthly, balances = aggregate_by_month(categorized_transactions)

    dds_data = {}

    sorted_months = sorted(monthly.keys(), key=lambda m: MONTH_ORDER.get(m, 99))

    for month in sorted_months:
        cats = monthly[month]
        month_data = {}

        # For each DDS line with a category key
        for level, label, cat_key, is_income in DDS_STRUCTURE:
            if cat_key is None:
                continue

            if cat_key == "total_abono":
                # Total sales = all abono
                month_data[cat_key] = cats["_total"]["abono"]
            elif cat_key == "opening_balance":
                month_data[cat_key] = balances.get(month, {}).get("opening", 0)
            elif cat_key == "closing_balance":
                month_data[cat_key] = balances.get(month, {}).get("closing", 0)
            elif cat_key == "_кв_бонус":
                # Квалификационный бонус — part of ОС лидерам, skip for now
                month_data[cat_key] = 0
            elif is_income:
                month_data[cat_key] = cats.get(cat_key, {}).get("abono", 0)
            else:
                month_data[cat_key] = cats.get(cat_key, {}).get("cargo", 0)

        dds_data[month] = month_data

    return dds_data


def write_dds_excel(dds_data, output_path=None):
    """Write DDS data to Excel file matching template format.

    Creates plan/fact columns per month.
    """
    output_path = output_path or os.path.join(OUTPUT_DIR, "ДДС_факт.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ДДС факт"

    # Styles
    header_font = Font(bold=True, size=12)
    section_font = Font(bold=True, size=11)
    subsection_font = Font(bold=True, size=10)
    number_format = '#,##0.00'
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    section_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    subtotal_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    sorted_months = sorted(dds_data.keys(), key=lambda m: MONTH_ORDER.get(m, 99))

    # Header row
    ws.cell(1, 1, "FL COSMETICS MEXICO — ДДС (факт)").font = header_font
    ws.cell(2, 1, "Статья ДДС").font = subsection_font
    ws.cell(2, 1).fill = header_fill
    ws.cell(2, 1).font = header_font_white

    for j, month in enumerate(sorted_months):
        col = 2 + j
        cell = ws.cell(2, col, month)
        cell.fill = header_fill
        cell.font = header_font_white
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    ws.column_dimensions['A'].width = 45

    # Data rows
    row_num = 3
    for level, label, cat_key, is_income in DDS_STRUCTURE:
        # Indent based on level
        indent = "  " * (level - 1)
        cell = ws.cell(row_num, 1, f"{indent}{label}")
        cell.border = thin_border

        if level <= 2:
            cell.font = section_font
            if level == 1:
                cell.fill = section_fill

        # Values per month
        if cat_key:
            for j, month in enumerate(sorted_months):
                col = 2 + j
                val = dds_data[month].get(cat_key, 0)
                value_cell = ws.cell(row_num, col, val)
                value_cell.number_format = number_format
                value_cell.border = thin_border
                value_cell.alignment = Alignment(horizontal='right')
        else:
            # Subtotal row — compute sum of children
            for j, month in enumerate(sorted_months):
                col = 2 + j
                ws.cell(row_num, col).border = thin_border

        row_num += 1

    # Add totals row
    row_num += 1
    ws.cell(row_num, 1, "ИТОГО приток").font = section_font
    for j, month in enumerate(sorted_months):
        col = 2 + j
        total_in = sum(
            dds_data[month].get(cat_key, 0)
            for _, _, cat_key, is_income in DDS_STRUCTURE
            if cat_key and is_income is True
        )
        ws.cell(row_num, col, total_in).number_format = number_format

    row_num += 1
    ws.cell(row_num, 1, "ИТОГО отток").font = section_font
    for j, month in enumerate(sorted_months):
        col = 2 + j
        total_out = sum(
            dds_data[month].get(cat_key, 0)
            for _, _, cat_key, is_income in DDS_STRUCTURE
            if cat_key and is_income is False
        )
        ws.cell(row_num, col, total_out).number_format = number_format

    row_num += 1
    ws.cell(row_num, 1, "Чистый денежный поток").font = Font(bold=True, size=11)
    for j, month in enumerate(sorted_months):
        col = 2 + j
        total_in = sum(
            dds_data[month].get(cat_key, 0)
            for _, _, cat_key, is_income in DDS_STRUCTURE
            if cat_key and is_income is True
        )
        total_out = sum(
            dds_data[month].get(cat_key, 0)
            for _, _, cat_key, is_income in DDS_STRUCTURE
            if cat_key and is_income is False
        )
        net = total_in - total_out
        cell = ws.cell(row_num, col, net)
        cell.number_format = number_format
        cell.font = Font(bold=True)

    wb.save(output_path)
    print(f"DDS saved to {output_path}")
    return output_path


def print_dds_summary(dds_data):
    """Print DDS summary to console."""
    sorted_months = sorted(dds_data.keys(), key=lambda m: MONTH_ORDER.get(m, 99))

    print(f"\n{'Статья ДДС':<40s}", end="")
    for month in sorted_months:
        short = month[:8]
        print(f"{short:>15s}", end="")
    print()
    print("-" * (40 + 15 * len(sorted_months)))

    for level, label, cat_key, is_income in DDS_STRUCTURE:
        indent = "  " * (level - 1)
        name = f"{indent}{label}"
        if cat_key:
            print(f"{name:<40s}", end="")
            for month in sorted_months:
                val = dds_data[month].get(cat_key, 0)
                if val:
                    print(f"{val:>15,.0f}", end="")
                else:
                    print(f"{'':>15s}", end="")
            print()
        elif level <= 2:
            print(f"\n{name}")

    # Net flow
    print(f"\n{'ЧИСТЫЙ ДЕНЕЖНЫЙ ПОТОК':<40s}", end="")
    for month in sorted_months:
        total_in = sum(
            dds_data[month].get(ck, 0)
            for _, _, ck, inc in DDS_STRUCTURE if ck and inc is True
        )
        total_out = sum(
            dds_data[month].get(ck, 0)
            for _, _, ck, inc in DDS_STRUCTURE if ck and inc is False
        )
        net = total_in - total_out
        print(f"{net:>15,.0f}", end="")
    print()


if __name__ == "__main__":
    from parser import parse_bank_statement
    from categorizer import (
        build_dictionary_from_bank, build_dictionary_from_william,
        merge_dictionaries, categorize_all
    )
    from config import PAYMENTS_FILE

    print("=" * 60)
    print("FL Cosmetics — ДДС Generator")
    print("=" * 60)

    # 1. Parse & categorize
    print("\n1. Parsing and categorizing transactions...")
    txns = parse_bank_statement()
    bank_dict = build_dictionary_from_bank(txns)
    william_dict = build_dictionary_from_william(PAYMENTS_FILE)
    merged = merge_dictionaries(bank_dict, william_dict)
    categorized = categorize_all(txns, merged)

    # 2. Generate DDS
    print("\n2. Generating DDS fact data...")
    dds_data = generate_dds_data(categorized)

    # 3. Print summary
    print_dds_summary(dds_data)

    # 4. Write Excel
    print("\n3. Writing Excel file...")
    path = write_dds_excel(dds_data)
    print(f"\nDone! File: {path}")
