"""
Category Dictionary Builder & Auto-Categorizer for FL Cosmetics.
Builds categorization rules from Emil's historical markup + William payments.

Key insight: For SPEI transactions (majority), the `concept` field contains
the counterparty name and payment purpose — this is the primary discriminator.
The `description` and `long_description` are generic ("PAGO TRANSFERENCIA SPEI").
"""

import json
import os
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from config import PAYMENTS_FILE, OUTPUT_DIR

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip3 install openpyxl")
    sys.exit(1)


# --- Normalize category names (Emil uses mixed case) ---

CATEGORY_NORMALIZATION = {
    "аренда склада": "Аренда склада",
    "бухгалтерские услуги": "Бухгалтерские услуги",
    "налог на выплаты лидерам": "Налог на выплаты лидерам",
    "налог на сотрудников": "Налог на сотрудников",
    "отправили из дп": "ОП",
    "стимулирование продаж": "Стимулирование продаж",
    "счет за интернет": "Счет за интернет",
    "транспортные в регион": "Транспортные в регион",
    "упаковка": "Упаковка",
    "услуги банка": "Услуги банка",
    "коммунальные расходы": "Коммунальные платежи",
    "хозрасходы": "Хозяйственные расходы",
    "услуги связи": "Связь",
    "счет за интернет": "Связь",
    "фот склад": "ЗП склад",
    "фот офис": "ЗП офис",
}


def normalize_category(cat):
    """Normalize category to canonical form."""
    if not cat:
        return ""
    cat = cat.strip()
    return CATEGORY_NORMALIZATION.get(cat.lower(), cat)


# --- Concept-based keyword rules ---
# These are patterns found in the `concept` field that reliably predict category.

CONCEPT_KEYWORD_RULES = [
    # ОС лидерам — leader payments (Asimilados + name, General + name)
    (r"(?i)^asimilados\b", "ОС лидерам"),
    (r"(?i)^general\b", "ОС лидерам"),

    # ЗП склад / ФОТ склад — warehouse salaries (ФОТ склад normalized to ЗП склад)
    # Named warehouse staff → always склад
    (r"(?i)^nomina\s+(carlos|bryan|pilar|rocio)\b", "ЗП склад"),
    (r"(?i)^reembolso\s+pago\s+de\s+(carlos|bryan|pilar|rocio)\b", "ЗП склад"),
    (r"(?i)^finiquito\b", "ЗП склад"),
    # Generic "Nomina REF" → split by amount in categorize_transaction()
    # Emil: ≤10,000 MXN → ЗП склад, >10,000 MXN → ЗП офис
    (r"(?i)^nomina\s+ref\b", "_nomina_split"),
    (r"(?i)^nomina$", "_nomina_split"),
    (r"(?i)^aguinaldo\b", "_nomina_split"),

    # ЗП офис — office salaries (Fondo, Anastasia)
    (r"(?i)^fondo\s+(anastasia|reembolsable)\b", "ЗП офис"),
    (r"(?i)^fondo\s+ref\b", "ЗП офис"),
    (r"(?i)^reembolso\s+fondo\b", "ЗП офис"),
    (r"(?i)^fondo\s+anastasia\b", "ЗП офис"),

    # Аренда склада
    (r"(?i)^renta\b", "Аренда склада"),

    # Бухгалтерские услуги
    (r"(?i)contabilidad\b", "Бухгалтерские услуги"),
    (r"(?i)^factura\s+contabilidad\b", "Бухгалтерские услуги"),
    (r"(?i)^asesoria\b", "Бухгалтерские услуги"),
    (r"(?i)^actualizacion\s+contrato\b", "Бухгалтерские услуги"),

    # Коммунальные платежи
    (r"(?i)\bcfe\b", "Коммунальные платежи"),
    (r"(?i)\bagua\b", "Коммунальные платежи"),
    (r"(?i)^factura\s+aqua\b", "Коммунальные платежи"),
    (r"(?i)^manejo\s+de\s+residuos\b", "Коммунальные платежи"),

    # Стимулирование продаж — events, marketing
    (r"(?i)^evento\b", "Стимулирование продаж"),
    (r"(?i)^factura\s+evento\b", "Стимулирование продаж"),
    (r"(?i)\bevento\b.*\bcancun\b", "Стимулирование продаж"),
    (r"(?i)^reunion\b", "Стимулирование продаж"),
    (r"(?i)^pago\s+marketing\b", "Стимулирование продаж"),
    (r"(?i)^segundo\s+pago\s+evento\b", "Стимулирование продаж"),

    # Транспортные в регион — Estafeta tracking numbers (961...)
    (r"^9[0-9]{9,}", "Транспортные в регион"),

    # Реклама каталог
    (r"(?i)catalog", "Реклама каталог"),

    # Связь / Интернет
    (r"(?i)\binternet\b", "Связь"),
    (r"(?i)\btelmex\b", "Связь"),

    # Упаковка
    (r"(?i)^caja\s+y\s+separador\b", "Упаковка"),
    (r"(?i)^bolsas\b", "Упаковка"),
    (r"(?i)\bbioempaques\b", "Расходные материалы"),

    # Услуги по таможенному оформлению
    (r"(?i)^certificaciones\b", "Услуги по таможенному оформлению"),
    (r"(?i)^impuest[or]s?\s+puerto\b", "Услуги по таможенному оформлению"),
    (r"(?i)^seguro\s+carga\b", "Услуги по таможенному оформлению"),

    # Хозяйственные расходы
    (r"(?i)^cambio\s+de\s+bateria\b", "Хозяйственные расходы"),
    (r"(?i)^cosmetiqueras\b", "Хозяйственные расходы"),
    (r"(?i)^camaras\b", "Хозяйственные расходы"),
    (r"(?i)^productos\s+de\s+limpieza\b", "Хозяйственные расходы"),
    (r"(?i)^instalacion\b", "Хозяйственные расходы"),
    (r"(?i)^reembolso\s+manguera\b", "Хозяйственные расходы"),

    # Взнос в УК
    (r"(?i)^prestamo\s+andrei\b", "Взнос в УК"),

    # Прочие расходы
    (r"(?i)^infonavit\b", "Прочие расходы"),
    (r"(?i)^baja\s+de\b", "Прочие расходы"),

    # Возврат средств
    (r"(?i)^reembolso\s+cuenta\s+faberlic\b", "Возврат средств"),
    (r"(?i)^reembolso\s+(pedido|producto|por pedido)\b", "Возврат средств"),

    # ОП — incoming payments (including MercadoPago identified by "MP" or "Mercado Pago" in concept)
    (r"(?i)^mercado\*?pago\b", "ОП"),
    (r"(?i)^mp\b", "ОП"),
    (r"(?i)^devolucion\b", "ОП"),
    (r"(?i)^dev\s+saf\b", "ОП"),
    (r"(?i)^pago\s+xyz\b", "ОП"),

    # ЗП офис — marketing/consulting services (Emil confirmed: Asesoria ventas = ЗП офис)
    (r"(?i)^servicios?\s+de\s+(marketing|asesoria|consultoria)\b", "ЗП офис"),
    (r"(?i)^asesoria\s+ventas\b", "ЗП офис"),
    (r"(?i)^pago\s+marketing\b", "ЗП офис"),
    (r"(?i)^comisiones\s+\w+", "ОС лидерам"),

    # Расходные материалы — packaging, labels
    (r"(?i)^bio\s*empaques\b", "Расходные материалы"),
    (r"(?i)^reembolso\s+pago\s+etiquetas\b", "Расходные материалы"),
    (r"(?i)^diferencia\s+etiquetas\b", "Расходные материалы"),

    # Стимулирование продаж — events, tours, momentum
    (r"(?i)\bgira\s+faberlic\b", "Стимулирование продаж"),
    (r"(?i)\bmomentum\b", "Стимулирование продаж"),
    (r"(?i)^liquidacion\b", "Стимулирование продаж"),
    (r"(?i)^bono\s+director\b", "Стимулирование продаж"),
    (r"(?i)^curso\b", "Стимулирование продаж"),

    # Хозяйственные расходы
    (r"(?i)^servicio\s+montacargas\b", "Хозяйственные расходы"),
    (r"(?i)^recoleccion\s+residuos\b", "Коммунальные платежи"),
    (r"(?i)^fumigacion\b", "Хозяйственные расходы"),

    # НДС — federal tax payments
    (r"(?i)^reembolso\s+factura\s+mercado\b", "Услуги банка"),
    (r"(?i)^facturacion\b", "Бухгалтерские услуги"),

    # Resico payments — leader tax regime
    (r"(?i)^resico\b", "Налог на выплаты лидерам"),

    # ОС лидерам — typos of Asimilados / General
    (r"(?i)^asimi[lo]lados\b", "ОС лидерам"),
    (r"(?i)^asimolados\b", "ОС лидерам"),
    (r"(?i)^asdimilados\b", "ОС лидерам"),
    (r"(?i)^asimilaldos\b", "ОС лидерам"),
    (r"(?i)^asimiladoss?\w*\b", "ОС лидерам"),
    (r"(?i)^generall\b", "ОС лидерам"),

    # Стимулирование продаж — events, travel, hospitality
    (r"(?i)^asamblea\b", "Стимулирование продаж"),
    (r"(?i)^adelanto\s+evento\b", "Стимулирование продаж"),
    (r"(?i)^anticipo\s+evento\b", "Стимулирование продаж"),
    (r"(?i)^segunda\s+parte\s+evento\b", "Стимулирование продаж"),
    (r"(?i)^desayuno\b", "Стимулирование продаж"),
    (r"(?i)^hospedaje", "Стимулирование продаж"),
    (r"(?i)^vuelos?\b", "Стимулирование продаж"),
    (r"(?i)^viaticos\b", "Стимулирование продаж"),
    (r"(?i)^boleto\b", "Стимулирование продаж"),
    (r"(?i)^flores\s+lideres\b", "Стимулирование продаж"),
    (r"(?i)^cierre\b", "Стимулирование продаж"),
    (r"(?i)^reserva\s+hotel\b", "Стимулирование продаж"),
    (r"(?i)^reembolso\s+(airbnb|viva|etn|aeromexico|hospadaje|hospedaje|vuelos)\b", "Стимулирование продаж"),
    (r"(?i)^reembolso\s+factura\s+viva", "Стимулирование продаж"),
    (r"(?i)^reembolso\s+de\s+gastos\b", "Стимулирование продаж"),

    # Транспортные — tracking numbers
    (r"^(260|987|GJ4)\d{6,}", "Транспортные в регион"),

    # Реклама каталог — print materials
    (r"(?i)^primer\s+pago\s+cat", "Реклама каталог"),
    (r"(?i)^brochure\b", "Реклама каталог"),
    (r"(?i)^anticipo\s+(brochure|cintas)\b", "Реклама каталог"),
    (r"(?i)^segundo\s+pago\s+cinta\b", "Реклама каталог"),
    (r"(?i)^agendas\b", "Реклама каталог"),
    (r"(?i)^logo\s+faberlic\b", "Реклама каталог"),
    (r"(?i)^rollos?\s+publicitarios\b", "Реклама каталог"),
    (r"(?i)^roll\s+ups?\b", "Реклама каталог"),
    (r"(?i)^maletines\b", "Реклама каталог"),
    (r"(?i)^papel\s+faberlic\b", "Расходные материалы"),

    # Emil: Carga = container shipment → goes to DDS only, PL uses Себестоимость
    (r"(?i)^carga\s+(enero|febrero|marzo|abril|mayo|junio|julio|agosto|sept|oct|nov|dic)", "Закупка товара (контейнер)"),
    # Emil: Turquia = стимулирование продаж
    (r"(?i)^turquia\b", "Стимулирование продаж"),
    (r"(?i)^boleto\s+turquia\b", "Стимулирование продаж"),
    # Emil: Impuestos puerto = НДС импорт → услуги по таможенному оформлению
    # (already handled by existing rule)
    # Suplementos = товар
    (r"(?i)^suplementos\b", "Закупка товара (контейнер)"),

    # ЗП офис — services, fond typos
    (r"(?i)^ventas\s+ref\b", "ЗП офис"),
    (r"(?i)^fodo\b", "ЗП офис"),

    # Коммунальные платежи
    (r"(?i)^manejo\s+residuos\b", "Коммунальные платежи"),
    (r"(?i)^residuos\b", "Коммунальные платежи"),
    (r"(?i)^suministro\s+a[gq]ua\b", "Коммунальные платежи"),

    # Хозяйственные расходы
    (r"(?i)^arreglo\b", "Хозяйственные расходы"),
    (r"(?i)^factura\s+arreglo\b", "Хозяйственные расходы"),
    (r"(?i)^extintores\b", "Хозяйственные расходы"),
    (r"(?i)^factura\s+productos\s+\d+\b", "Хозяйственные расходы"),

    # Финансы
    (r"(?i)^pago\s+prestamo\b", "Взнос в УК"),
    (r"(?i)^deposito\s+en\s+garantia\b", "Аренда склада"),
    (r"(?i)^endoso\b", "Прочие расходы"),

    # Прочие
    (r"(?i)^seguro\s+ref\b", "Прочие расходы"),
    (r"(?i)^actualizacion\s+inm\b", "Прочие расходы"),
    (r"(?i)^reembolso\s+regalos\b", "Стимулирование продаж"),
    (r"(?i)^apertura\s+cuent", "Прочие расходы"),
    (r"(?i)^segunda\s+audiencia\b", "Прочие расходы"),
    (r"(?i)^diferencia\s+fondo\b", "ЗП офис"),
    (r"(?i)^reembolso\s+cuenta\s+fl\b", "Возврат средств"),
    (r"(?i)^reembolso\s+por\s+producto\b", "Возврат средств"),

    # Бухгалтерские услуги — invoices
    (r"(?i)^factura\s+(noms|5193|septiembre)\b", "Бухгалтерские услуги"),
    # Emil: Complementario = печать каталога → Реклама каталог
    (r"(?i)^pago\s+complementario\b", "Реклама каталог"),
    # Emil: Facturacion = JT → Транспортные в регион
    (r"(?i)^facturacion\b", "Транспортные в регион"),

    # Last remaining patterns
    (r"(?i)^reembolso\s+factura\s+(luz|cfe|estrella|officemax)", "Коммунальные платежи"),
    (r"(?i)^servicios\s+ventas\b", "ЗП офис"),
]

# Description-level catch-all rules
DESCRIPTION_RULES_EXTRA = [
    (r"(?i)^ab\s+dev\s+tra\s+int\b", "ОП"),
    (r"(?i)^com\s+rep\s+token\b", "Услуги банка"),
]

# Description-based rules (for non-SPEI transactions)
DESCRIPTION_RULES = [
    (r"(?i)^comision\b", "Услуги банка"),
    (r"(?i)^membresia\b", "Услуги банка"),
    (r"(?i)^cargo\s+comision\b", "Услуги банка"),
    (r"(?i)^renta\s+tpv\b", "Услуги банка"),
    (r"(?i)^iva\s+(x\s+)?comision\b", "Услуги банка"),
    (r"(?i)^com\s+trans\s+int\b", "Услуги банка"),
    (r"(?i)^iva\s+comision\b", "Услуги банка"),
    (r"(?i)^pago\s+imp", "НДС"),
    (r"(?i)^cgo\s+imp\s+fedtra\b", "НДС"),
    (r"(?i)^ab\s+transf\s+spei\b", "ОП"),
    (r"(?i)^abono\s+por\s+devol", "ОП"),
    (r"(?i)^abono\s+transferencia\s+spei\b", "ОП"),
    (r"(?i)^deposito\s+efectivo", "ОП"),
    (r"(?i)^pago\s+cheque\b", "ОС лидерам"),
    (r"(?i)^cgo\s+trans\s+(elec|inte)\b", "ЗП офис"),
    (r"(?i)^compensa\s+spei\b", "_skip"),
]


# --- William payments category mapping ---

WILLIAM_CATEGORY_MAP = {
    "зарплата": "ЗП склад",
    "зп": "ЗП склад",
    "зп фернанде": "ЗП офис",
    "зп майре": "ЗП офис",
    "зп сесилии": "ЗП склад",
    "зп карлосу": "ЗП склад",
    "зп пилар": "ЗП склад",
    "зп брайану": "ЗП склад",
    "зп росио": "ЗП склад",
    "фонд анастасия": "ОС лидерам",
    "предоплата за каталоги": "Реклама каталог",
    "за каталоги": "Реклама каталог",
    "эстафета": "Транспортные в регион",
    "эстафета декабрь": "Транспортные в регион",
    "эстафета январь": "Транспортные в регион",
    "эстафета февраль": "Транспортные в регион",
    "эстафета март": "Транспортные в регион",
    "dhl": "Транспортные в регион",
    "счет за свет": "Коммунальные платежи",
    "за интернет": "Связь",
    "за воду": "Коммунальные платежи",
    "за свет": "Коммунальные платежи",
    "аренда": "Аренда склада",
    "аренда склада": "Аренда склада",
    "за травление насекомых": "Хозяйственные расходы",
    "за вывоз мусора": "Хозяйственные расходы",
    "за запчасти для стеллажей": "Хозяйственные расходы",
    "упаковочный материал": "Упаковка",
    "за упаковку": "Упаковка",
    "маркетинг": "Стимулирование продаж",
    "маркетинговые услуги": "Стимулирование продаж",
}


def _map_william_category(russian_desc):
    """Map William's Russian description to DDS category."""
    if not russian_desc:
        return None
    desc_lower = russian_desc.strip().lower()

    # Direct match
    if desc_lower in WILLIAM_CATEGORY_MAP:
        return WILLIAM_CATEGORY_MAP[desc_lower]

    # Partial match
    for key, cat in WILLIAM_CATEGORY_MAP.items():
        if key in desc_lower:
            return cat

    return None


# --- Dictionary builders ---


def build_dictionary_from_bank(transactions):
    """Build category dictionary from Emil's historical markup.

    Creates frequency-based mappings:
    - by_concept: concept text → {category: count} (primary for SPEI)
    - by_description: description → {category: count} (for non-SPEI)
    - by_counterparty: counterparty name → {category: count} (from William)
    """
    by_concept = defaultdict(Counter)
    by_description = defaultdict(Counter)
    by_counterparty = defaultdict(Counter)

    for txn in transactions:
        cat = normalize_category(txn.get("dds_category", ""))
        if not cat:
            continue

        desc = txn.get("description", "").strip()
        concept = txn.get("concept", "").strip()

        # Map concept → category (primary discriminator)
        if concept:
            by_concept[concept][cat] += 1
            # Also extract counterparty from concept
            # "Asimilados Gonzalez Gomez Carla REF 0000000" → extract name
            concept_clean = re.sub(r'\s+REF\s+\d+$', '', concept).strip()
            if concept_clean != concept.strip():
                by_concept[concept_clean][cat] += 1

        # Map description → category
        if desc:
            by_description[desc][cat] += 1

    return {
        "by_concept": {k: dict(v) for k, v in by_concept.items()},
        "by_description": {k: dict(v) for k, v in by_description.items()},
        "by_counterparty": {k: dict(v) for k, v in by_counterparty.items()},
    }


def build_dictionary_from_william(filepath=None):
    """Build category dictionary from William payments file.

    Reads 'Оплаты William .xlsx' — extracts counterparty → category mapping.
    """
    filepath = filepath or PAYMENTS_FILE

    if not os.path.exists(filepath):
        print(f"WARNING: William payments file not found: {filepath}")
        return {"by_concept": {}, "by_description": {}, "by_counterparty": {}}

    wb = openpyxl.load_workbook(filepath, data_only=True)
    by_counterparty = defaultdict(Counter)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            if row is None or len(row) < 6:
                continue

            counterparty = row[2] if len(row) > 2 else None
            russian_desc = row[5] if len(row) > 5 else None

            if not counterparty or not russian_desc:
                continue

            counterparty_str = str(counterparty).strip()
            russian_str = str(russian_desc).strip()

            if not counterparty_str or not russian_str:
                continue

            dds_cat = _map_william_category(russian_str)
            if dds_cat:
                by_counterparty[counterparty_str.upper()][dds_cat] += 1

    wb.close()

    return {
        "by_concept": {},
        "by_description": {},
        "by_counterparty": {k: dict(v) for k, v in by_counterparty.items()},
    }


def merge_dictionaries(bank_dict, william_dict):
    """Merge bank and William dictionaries.

    Saves merged dictionary to output/category_dictionary.json.
    Returns merged dict.
    """
    merged = {
        "by_concept": dict(bank_dict.get("by_concept", {})),
        "by_description": dict(bank_dict.get("by_description", {})),
        "by_counterparty": dict(bank_dict.get("by_counterparty", {})),
    }

    # Merge William counterparties
    for cp, cats in william_dict.get("by_counterparty", {}).items():
        if cp not in merged["by_counterparty"]:
            merged["by_counterparty"][cp] = cats
        else:
            existing = merged["by_counterparty"][cp]
            for cat, count in cats.items():
                existing[cat] = existing.get(cat, 0) + count

    # Save to JSON
    output_path = os.path.join(OUTPUT_DIR, "category_dictionary.json")
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(merged, f, ensure_ascii=False, indent=2)
    print(f"Dictionary saved to {output_path}")

    return merged


def _best_category(counter_dict):
    """Get the most frequent category from a {category: count} dict."""
    if not counter_dict:
        return None, 0
    best = max(counter_dict, key=counter_dict.get)
    total = sum(counter_dict.values())
    return best, counter_dict[best] / total if total > 0 else 0


def categorize_transaction(txn, dictionary):
    """Categorize a single transaction using rules + dictionary.

    Matching priority:
    1. Concept keyword rules (regex on concept) → 0.95
    2. Exact concept match in dictionary → 0.90
    3. Concept substring match in dictionary → 0.85
    4. Description rules (regex) → 0.80
    5. Description dictionary match → 0.75
    6. Concept keyword match in concept → 0.70
    7. No match → ("UNKNOWN", 0.0)

    Returns (category_name, confidence_score).
    """
    desc = txn.get("description", "").strip()
    concept = txn.get("concept", "").strip()
    comment = txn.get("comment", "").strip()

    by_concept = dictionary.get("by_concept", {})
    by_desc = dictionary.get("by_description", {})
    by_cp = dictionary.get("by_counterparty", {})

    # 0. If Emil already wrote the category in comment field (same as dds_category)
    # — skip, we evaluate against dds_category

    # 1. Concept keyword rules (most reliable)
    if concept:
        for pattern, cat in CONCEPT_KEYWORD_RULES:
            if re.search(pattern, concept):
                # Handle special markers
                if cat == "_nomina_split":
                    # Emil: ≤10K → ЗП склад, >10K → ЗП офис
                    cargo = txn.get("cargo", 0) or 0
                    return ("ЗП склад" if cargo <= 10000 else "ЗП офис"), 0.90
                if cat == "_skip":
                    return "Внутренний перевод", 0.50
                return cat, 0.95

    # 2. Exact concept match in dictionary
    if concept:
        concept_clean = re.sub(r'\s+REF\s+\d+$', '', concept).strip()

        if concept in by_concept:
            cat, ratio = _best_category(by_concept[concept])
            if cat and ratio >= 0.6:
                return cat, 0.90

        if concept_clean != concept and concept_clean in by_concept:
            cat, ratio = _best_category(by_concept[concept_clean])
            if cat and ratio >= 0.6:
                return cat, 0.90

    # 3. Counterparty match — check if concept contains a known counterparty
    if concept:
        concept_upper = concept.upper()
        best_cp_match = None
        best_cp_len = 0
        for cp, cats in by_cp.items():
            if len(cp) >= 4 and cp in concept_upper and len(cp) > best_cp_len:
                cat, ratio = _best_category(cats)
                if cat:
                    best_cp_match = cat
                    best_cp_len = len(cp)
        if best_cp_match:
            return best_cp_match, 0.85

    # 4. Description rules
    if desc:
        for pattern, cat in DESCRIPTION_RULES:
            if re.search(pattern, desc):
                if cat == "_skip":
                    return "Внутренний перевод", 0.50
                return cat, 0.80
        for pattern, cat in DESCRIPTION_RULES_EXTRA:
            if re.search(pattern, desc):
                return cat, 0.75

    # 5. Non-generic description match in dictionary
    # Skip generic descriptions that map to many categories
    generic_descriptions = {
        "PAGO TRANSFERENCIA SPEI", "ABONO TRANSFERENCIA SPEI",
        "PAGO TRAN SPEI", "AB TRANSF SPEI",
        "TRANSF INTERNACIONAL ENVIADA", "ABONO POR DEVOLUCIONES SPEI",
    }
    if desc and desc not in generic_descriptions and desc in by_desc:
        cat, ratio = _best_category(by_desc[desc])
        if cat and ratio >= 0.5:
            return cat, 0.75

    # 6. Concept substring matching against dictionary concepts
    if concept:
        concept_upper = concept.upper()
        for known_concept, cats in by_concept.items():
            known_upper = known_concept.upper()
            # Only match if significant overlap
            if len(known_upper) >= 8 and known_upper in concept_upper:
                cat, ratio = _best_category(cats)
                if cat and ratio >= 0.7:
                    return cat, 0.70

    return "UNKNOWN", 0.0


def categorize_all(transactions, dictionary):
    """Run categorize_transaction on all transactions.

    Adds fields: predicted_category, confidence, is_correct.
    Returns enriched list of dicts.
    """
    result = []
    for txn in transactions:
        predicted, confidence = categorize_transaction(txn, dictionary)
        enriched = dict(txn)
        enriched["predicted_category"] = predicted
        enriched["confidence"] = confidence

        actual = normalize_category(txn.get("dds_category", ""))
        if actual:
            enriched["is_correct"] = (predicted == actual)
        else:
            enriched["is_correct"] = None

        result.append(enriched)

    return result


def calculate_accuracy(categorized):
    """Calculate accuracy on transactions that have both predicted and actual categories.

    Returns dict: total, correct, accuracy_pct, unknown_count, mismatches (list).
    """
    with_actual = [t for t in categorized if t.get("is_correct") is not None]

    total = len(with_actual)
    correct = sum(1 for t in with_actual if t["is_correct"])
    unknown = sum(1 for t in with_actual if t["predicted_category"] == "UNKNOWN")

    mismatches = []
    for t in with_actual:
        if not t["is_correct"] and t["predicted_category"] != "UNKNOWN":
            mismatches.append({
                "description": t.get("description", ""),
                "concept": t.get("concept", "")[:60],
                "actual": normalize_category(t.get("dds_category", "")),
                "predicted": t["predicted_category"],
                "confidence": t["confidence"],
            })

    return {
        "total": total,
        "correct": correct,
        "accuracy_pct": (correct / total * 100) if total > 0 else 0.0,
        "unknown_count": unknown,
        "mismatches": mismatches,
    }


if __name__ == "__main__":
    from parser import parse_bank_statement

    print("=" * 60)
    print("FL Cosmetics — Category Dictionary Builder")
    print("=" * 60)

    # 1. Parse bank statement
    print("\n1. Parsing bank statement...")
    txns = parse_bank_statement()
    print(f"   Total transactions: {len(txns)}")
    with_cat = sum(1 for t in txns if t.get("dds_category"))
    print(f"   With DDS category: {with_cat}")

    # 2. Build dictionary from bank markup
    print("\n2. Building dictionary from bank markup...")
    bank_dict = build_dictionary_from_bank(txns)
    print(f"   Concepts: {len(bank_dict['by_concept'])}")
    print(f"   Descriptions: {len(bank_dict['by_description'])}")

    # 3. Build dictionary from William payments
    print("\n3. Building dictionary from William payments...")
    william_dict = build_dictionary_from_william(PAYMENTS_FILE)
    print(f"   Counterparties: {len(william_dict['by_counterparty'])}")

    # 4. Merge
    print("\n4. Merging dictionaries...")
    merged = merge_dictionaries(bank_dict, william_dict)
    print(f"   Total counterparties: {len(merged['by_counterparty'])}")
    print(f"   Total concepts: {len(merged['by_concept'])}")
    print(f"   Total descriptions: {len(merged['by_description'])}")

    # 5. Categorize
    print("\n5. Running auto-categorization...")
    categorized = categorize_all(txns, merged)

    # 6. Accuracy
    metrics = calculate_accuracy(categorized)
    print(f"\n{'=' * 60}")
    print(f"ACCURACY REPORT")
    print(f"{'=' * 60}")
    print(f"   Transactions with actual category: {metrics['total']}")
    print(f"   Correctly predicted: {metrics['correct']}")
    print(f"   Accuracy: {metrics['accuracy_pct']:.1f}%")
    print(f"   Unknown (no prediction): {metrics['unknown_count']}")
    print(f"   Wrong predictions: {len(metrics['mismatches'])}")

    # 7. Top mismatches
    if metrics["mismatches"]:
        print(f"\n   Top mismatches:")
        for m in metrics["mismatches"][:20]:
            print(f"     concept={m['concept'][:40]:40s} actual={m['actual']:25s} pred={m['predicted']:25s} ({m['confidence']:.2f})")

    # 8. Accuracy by category
    print(f"\n   Accuracy by category:")
    from collections import defaultdict
    cat_stats = defaultdict(lambda: {"total": 0, "correct": 0})
    for t in categorized:
        if t.get("is_correct") is not None:
            actual = normalize_category(t.get("dds_category", ""))
            cat_stats[actual]["total"] += 1
            if t["is_correct"]:
                cat_stats[actual]["correct"] += 1
    for cat in sorted(cat_stats.keys()):
        s = cat_stats[cat]
        pct = s["correct"] / s["total"] * 100 if s["total"] > 0 else 0
        marker = " !!!" if pct < 50 else ""
        print(f"     {cat:35s} {s['correct']:3d}/{s['total']:3d} = {pct:5.1f}%{marker}")

    # 9. Unknown transactions
    unknown_txns = [t for t in categorized if t["predicted_category"] == "UNKNOWN" and not t.get("dds_category")]
    print(f"\n   Uncategorized (no actual, no prediction): {len(unknown_txns)}")
    if unknown_txns:
        print(f"   Top 10 unknown:")
        for t in unknown_txns[:10]:
            cargo = t.get("cargo", 0)
            abono = t.get("abono", 0)
            amt = f"C:{cargo:,.0f}" if cargo else f"A:{abono:,.0f}"
            print(f"     [{t['month'][:10]}] {t['description'][:25]:25s} concept={t.get('concept','')[:40]:40s} {amt}")

    print(f"\nDone.")
